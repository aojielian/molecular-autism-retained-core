
from pathlib import Path
import os
import re
import gzip
import csv
import json
import subprocess
import tempfile
from typing import List, Dict, Optional, Iterable, Set
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(os.environ.get('MA_REVISION_ROOT', '/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/MolecularAutism_Revision'))
OUTDIR = ROOT / 'Supplementary_Tables_v4'
OUTFILE = OUTDIR / 'Supplementary_Tables_Master_v4.xlsx'
OUTDIR.mkdir(parents=True, exist_ok=True)

# -------------------------------------------------
# Manual values that are not always recoverable from exported TSVs
# -------------------------------------------------
MANUAL_S10_VALUES = {
    'discovery_clustering_resolution': '0.4',
    'discovery_reclustering_rationale': 'Selected to preserve the dominant candidate/reference contrast while avoiding over-interpretation of finer fragmentation.',
    'validation_reclustering_resolution': '0.4',
    'validation_reclustering_rationale': 'Default val_resolution used in the finalized Velmeshev score-based validation scripts.',
    'validation_npcs': '20',
    'candidate_quantile': '0.75',
    'reference_quantile': '0.25',
    'mapping_corr_delta_threshold': '0.05',
    'mapping_state_delta_threshold': '0.00',
    'min_cells_sample': '20',
    'min_cells_group': '10',
    'bulk_model_covariates': 'region, batch, sex, ancestry, age, age_squared, postmortem_interval, RNA_integrity',
    'github_repository': os.environ.get('MA_GITHUB_REPOSITORY', 'https://github.com/aojielian/molecular-autism-retained-core'),
    'zenodo_archive': os.environ.get('MA_ZENODO_ARCHIVE', 'REPLACE_WITH_FINAL_ZENODO_DOI'),
}

PACKAGE_PURPOSE = {
    'Seurat': 'single-cell preprocessing, clustering, UMAP, plotting',
    'SeuratObject': 'Seurat object infrastructure',
    'harmony': 'donor-aware integration / embedding harmonization',
    'ggplot2': 'data visualization',
    'patchwork': 'multi-panel figure assembly',
    'data.table': 'tabular data handling',
    'dplyr': 'data manipulation',
    'Matrix': 'sparse matrix operations',
    'edgeR': 'pseudobulk differential expression',
    'limma': 'linear modeling and voom workflow',
    'nlme': 'linear/mixed modeling',
    'lme4': 'linear/mixed modeling',
    'AnnotationDbi': 'gene identifier annotation',
    'org.Hs.eg.db': 'human gene annotation mapping',
    'fgsea': 'gene-set enrichment analysis',
    'clusterProfiler': 'functional enrichment / annotation support',
    'optparse': 'command-line argument parsing',
    'statmod': 'supplementary statistical modeling support',
}

CURATED_PACKAGES = [
    'Seurat', 'SeuratObject', 'harmony', 'ggplot2', 'patchwork', 'data.table',
    'dplyr', 'Matrix', 'edgeR', 'limma', 'nlme', 'lme4', 'AnnotationDbi',
    'org.Hs.eg.db', 'fgsea', 'clusterProfiler', 'optparse', 'statmod'
]

COHORT_META = {
    'Discovery': {
        'kind': 'tsv',
        'file': Path('/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/PsychENCODE_Science_2024/meta.tsv'),
        'cell_id': 'Cell_ID',
        'sample_id': 'orig.ident',
        'subject_id': 'individual_ID',
        'diagnosis': 'Diagnosis',
        'age': 'Age',
        'sex': 'Sex_Chromosome',
        'brain_region': 'Brain_Region',
        'platform': 'Chemistry10X',
        'platform_manual': '',
    },
    'Velmeshev validation': {
        'kind': 'tsv',
        'file': Path('/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/Velmeshev_scRNA/meta.tsv'),
        'cell_id': 'cell',
        'sample_id': 'sample',
        'subject_id': 'individual',
        'diagnosis': 'diagnosis',
        'age': 'age',
        'sex': 'sex',
        'brain_region': 'region',
        'platform': '',
        'platform_manual': os.environ.get('MA_VEL_PLATFORM', 'single-nucleus RNA-seq'),
    },
    'Gandal 2022 bulk cortex': {
        'kind': 'rdata',
        'file': Path('/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/Gandal_2022/Gene_NormalizedExpression_Metadata_wModelMatrix.RData'),
        'object_name': 'datMeta',
        'sample_id': 'sample_id',
        'subject_id': 'subject',
        'diagnosis': 'Diagnosis',
        'age': 'Age',
        'sex': 'Sex',
        'brain_region': 'region',
        'platform': 'SeqMethod',
        'read_length': 'Read_Length',
        'platform_manual': 'bulk RNA-seq',
    },
}

# -------------------------------------------------
# Helpers
# -------------------------------------------------

def read_tsv(path: Path) -> List[Dict[str, str]]:
    if not path or not path.exists():
        return []
    opener = gzip.open if path.suffix == '.gz' else open
    with opener(path, 'rt', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f, delimiter='\t')
        return [dict(row) for row in reader]


def read_txt_as_rows(path: Path, value_col: str, extra: Optional[Dict[str, str]] = None) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    extra = extra or {}
    rows = []
    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.rstrip('\n')
            if line.strip() == '':
                continue
            row = {value_col: line}
            row.update(extra)
            rows.append(row)
    return rows


def read_rdata_frame(path: Path, object_name: str) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    tmp = tempfile.NamedTemporaryFile(prefix='datMeta_', suffix='.tsv', delete=False)
    tmp_path = Path(tmp.name)
    tmp.close()
    r_code = f"""
    e <- new.env()
    load({json.dumps(str(path))}, envir = e)
    if (!exists({json.dumps(object_name)}, envir = e)) {{
      stop('Object not found in RData: {object_name}')
    }}
    obj <- get({json.dumps(object_name)}, envir = e)
    if (!is.data.frame(obj)) {{
      obj <- as.data.frame(obj, stringsAsFactors = FALSE)
    }}
    write.table(obj, file = {json.dumps(str(tmp_path))}, sep = "\\t", quote = FALSE, row.names = FALSE, col.names = TRUE, na = "")
    """
    try:
        subprocess.run(['Rscript', '-e', r_code], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        rows = read_tsv(tmp_path)
    except Exception as e:
        print(f'WARNING: failed to read {object_name} from {path}: {e}')
        rows = []
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
    return rows


def maybe_int(x):
    try:
        if x is None or x == '':
            return None
        if isinstance(x, (int, float)):
            return int(x)
        return int(float(str(x)))
    except Exception:
        return None


def maybe_float(x):
    try:
        if x is None or x == '':
            return None
        return float(str(x))
    except Exception:
        return None


def normalize_str(x):
    if x is None:
        return ''
    return str(x).strip()


def clean_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    out = []
    for r in rows:
        rr = {}
        for k, v in r.items():
            k2 = normalize_str(k)
            if k2 == '' or k2 == 'source_file':
                continue
            rr[k2] = '' if v is None else str(v).strip()
        if not rr:
            continue
        if all(v == '' for v in rr.values()):
            continue
        out.append(rr)
    return out


def add_section(rows: List[Dict[str, str]], section: str) -> List[Dict[str, str]]:
    out = []
    for r in clean_rows(rows):
        rr = {'section': section}
        rr.update(r)
        out.append(rr)
    return out


def combine_rows(parts: List[List[Dict[str, str]]]) -> List[Dict[str, str]]:
    out = []
    for p in parts:
        out.extend(clean_rows(p))
    return out


def unique_values(rows: List[Dict[str, str]], candidates: List[str]) -> List[str]:
    cols = set()
    for r in rows:
        cols.update(r.keys())
    return [c for c in candidates if c in cols]


def write_sheet(ws, rows: List[Dict[str, str]]):
    rows = clean_rows(rows)
    if not rows:
        ws['A1'] = 'note'
        ws['A2'] = f'No data found for {ws.title}'
        ws.freeze_panes = 'A2'
        return

    cols = []
    seen = set()
    preferred = [
        'Category', 'dataset', 'summary_scope', 'analysis_unit', 'section', 'analysis_stage', 'cluster',
        'n_ASD_cases', 'n_Control_cases', 'n_total_cases', 'qc_retained_subjects', 'qc_retained_nuclei',
        'qc_retained_samples', 'age_range_years', 'sex_distribution', 'sampled_brain_regions',
        'sequencing_platform', 'Item', 'Version_or_value', 'Purpose_or_definition', 'Notes', 'notes'
    ]
    for col in preferred:
        if any(col in r for r in rows) and col not in seen:
            cols.append(col)
            seen.add(col)

    for r in rows:
        for k in r.keys():
            if k not in seen:
                cols.append(k)
                seen.add(k)

    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F4E78')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='D9E2F3')
        c.border = Border(bottom=thin)

    for i, r in enumerate(rows, start=2):
        for j, col in enumerate(cols, start=1):
            ws.cell(row=i, column=j, value=r.get(col, ''))

    ws.freeze_panes = 'A2'

    for j, col in enumerate(cols, start=1):
        max_len = len(str(col))
        for i in range(2, min(len(rows) + 2, 500)):
            val = ws.cell(row=i, column=j).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 14), 56)

    ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
    table = Table(displayName=f"Tbl_{ws.title[:20].replace('-', '_')}", ref=ref)
    style = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def find_first_existing(candidates: Iterable[Path], glob_name: Optional[str] = None) -> Optional[Path]:
    for p in candidates:
        if p.exists():
            return p
    if glob_name:
        matches = list(ROOT.rglob(glob_name))
        if matches:
            return matches[0]
    return None


def read_first_existing_tsv(candidates: Iterable[Path], glob_name: Optional[str] = None) -> List[Dict[str, str]]:
    p = find_first_existing(candidates, glob_name=glob_name)
    return read_tsv(p) if p else []


def rows_to_index(rows: List[Dict[str, str]], key: str) -> Dict[str, Dict[str, str]]:
    idx = {}
    for r in clean_rows(rows):
        k = normalize_str(r.get(key))
        if k != '':
            idx[k] = r
    return idx


def infer_background_reason(marker_preview: str) -> Optional[str]:
    s = (marker_preview or '').lower()
    if s == '':
        return None

    neuronal = ['cntnap2', 'ptprd', 'ryr2', 'camk2a', 'sv2b', 'olfm1', 'gabrb2', 'nptxr', 'st18', 'il1rapl1', 'nrg1']
    astro = ['slc1a2', 'aqp4', 'gpc5', 'ryr3']
    oligo = ['mobp', 'tf', 'slc44a1', 'opalin', 'plp1', 'mbp']
    vascular = ['colec12', 'vcan', 'f13a1', 'cr1', 'selp']
    lymphoid = ['blnk', 'dock8', 'cd83']
    stress = ['hsph1', 'dnaja4', 'bag3', 'hspd1', 'serpine1']

    if any(x in s for x in neuronal):
        return 'neuronal_or_synaptic_background'
    if any(x in s for x in astro):
        return 'astrocytic_background'
    if any(x in s for x in oligo):
        return 'oligodendroglial_or_myelin_background'
    if any(x in s for x in vascular):
        return 'vascular_or_perivascular_background'
    if any(x in s for x in lymphoid):
        return 'lymphoid_or_immune_background'
    if any(x in s for x in stress):
        return 'stress_or_uncertain_identity'
    return None


def infer_primary_reason(row: Dict[str, str], retained_clusters: set) -> str:
    cluster = normalize_str(row.get('cluster'))
    if cluster in retained_clusters:
        return 'retained_core_microglial_cluster'

    top_nonmg = maybe_float(row.get('top_nonmg_score'))
    mg_total = maybe_float(row.get('microglia_total_score'))
    if top_nonmg is not None and mg_total is not None and top_nonmg > mg_total:
        return 'non_microglial_background'

    reason_from_markers = infer_background_reason(row.get('top_marker_preview') or row.get('top50_preview'))
    if reason_from_markers:
        return reason_from_markers

    nh = maybe_int(row.get('n_homeostatic_markers_in_top50'))
    na = maybe_int(row.get('n_activation_markers_in_top50'))
    if nh == 0 and na == 0:
        return 'weak_microglial_identity'

    n_donors = maybe_int(row.get('n_donors'))
    if n_donors is not None and n_donors < 40:
        return 'limited_donor_support'

    return 'not_retained_after_audit'


def parse_session_environment(session_file: Path) -> Dict[str, str]:
    env = {}
    if not session_file or not session_file.exists():
        return env
    with open(session_file, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            s = line.strip()
            if s.startswith('R version'):
                env['R'] = s.replace('R version', '').strip()
            elif s.startswith('Platform:'):
                env['Platform'] = s.replace('Platform:', '').strip()
            elif s.startswith('Running under:'):
                env['Operating system'] = s.replace('Running under:', '').strip()
    return env


def parse_package_versions(session_files: List[Path]) -> Dict[str, str]:
    pkg_map = {}
    token_re = re.compile(r'([A-Za-z][A-Za-z0-9\.]+)_([0-9][A-Za-z0-9\.\-]+)')
    for f in session_files:
        if not f.exists():
            continue
        with open(f, 'r', encoding='utf-8', errors='ignore') as fh:
            for line in fh:
                for name, ver in token_re.findall(line):
                    if name not in pkg_map:
                        pkg_map[name] = ver
    return pkg_map


def first_metric_counts(rows: List[Dict[str, str]]) -> Dict[str, Optional[int]]:
    rows = clean_rows(rows)
    for r in rows:
        n_asd = maybe_int(r.get('n_ASD'))
        n_ctl = maybe_int(r.get('n_Control'))
        if n_asd is not None and n_ctl is not None:
            return {'n_ASD': n_asd, 'n_Control': n_ctl}
    return {'n_ASD': None, 'n_Control': None}


def canonical_dx(x: str) -> str:
    s = normalize_str(x)
    sl = s.lower()
    if sl in {'ctl', 'control', 'no known disorder'}:
        return 'Control'
    if sl in {'asd', 'autism', 'case'} or sl.startswith('autism'):
        return 'ASD'
    if 'dup15q' in sl or 'duplication syndrome' in sl:
        return 'Dup15q'
    return s


def canonical_sex(x: str) -> str:
    s = normalize_str(x)
    sl = s.lower()
    if sl in {'xx', 'f', 'female'}:
        return 'F'
    if sl in {'xy', 'm', 'male'}:
        return 'M'
    return s


def extract_id_set(rows: List[Dict[str, str]], candidates: List[str]) -> Set[str]:
    out = set()
    cols = unique_values(rows, candidates)
    if not cols:
        return out
    col = cols[0]
    for r in rows:
        v = normalize_str(r.get(col))
        if v != '':
            out.add(v)
    return out


def summarize_age_range(values: List[str]) -> str:
    nums = sorted({round(v, 3) for v in [maybe_float(x) for x in values] if v is not None})
    if not nums:
        return ''
    lo = int(nums[0]) if float(nums[0]).is_integer() else nums[0]
    hi = int(nums[-1]) if float(nums[-1]).is_integer() else nums[-1]
    return f'{lo}-{hi}'


def summarize_distribution(values: List[str]) -> str:
    counts: Dict[str, int] = {}
    for v in values:
        v = normalize_str(v)
        if v == '':
            continue
        counts[v] = counts.get(v, 0) + 1
    if not counts:
        return ''
    parts = [f'{k}={counts[k]}' for k in sorted(counts.keys())]
    return '; '.join(parts)


def summarize_unique(values: List[str], sep: str = '; ') -> str:
    vals = sorted({normalize_str(v) for v in values if normalize_str(v) != ''})
    return sep.join(vals)


def normalize_platform_value(v: str) -> str:
    s = normalize_str(v)
    if s == '':
        return ''
    sl = s.lower()
    if sl == 'v2':
        return '10x Genomics Chromium V2'
    if sl == 'v3':
        return '10x Genomics Chromium V3'
    if sl == 'unstranded':
        return 'bulk RNA-seq (unstranded)'
    return s


def build_subject_records(meta_rows: List[Dict[str, str]], mapping: Dict[str, str]) -> Dict[str, Dict[str, str]]:
    subject_col = mapping.get('subject_id', '')
    sample_col = mapping.get('sample_id', '')
    dx_col = mapping.get('diagnosis', '')
    age_col = mapping.get('age', '')
    sex_col = mapping.get('sex', '')
    region_col = mapping.get('brain_region', '')
    platform_col = mapping.get('platform', '')
    read_length_col = mapping.get('read_length', '')

    subjects: Dict[str, Dict[str, str]] = {}
    for r in meta_rows:
        sid = normalize_str(r.get(subject_col)) if subject_col else ''
        if sid == '':
            sid = normalize_str(r.get(sample_col)) if sample_col else ''
        if sid == '':
            continue
        rec = subjects.setdefault(sid, {
            'diagnosis': '',
            'age': '',
            'sex': '',
            'regions': set(),
            'platforms': set(),
            'read_lengths': set(),
            'samples': set(),
        })
        if dx_col and rec['diagnosis'] == '':
            rec['diagnosis'] = canonical_dx(r.get(dx_col))
        if age_col and rec['age'] == '':
            rec['age'] = normalize_str(r.get(age_col))
        if sex_col and rec['sex'] == '':
            rec['sex'] = canonical_sex(r.get(sex_col))
        if region_col:
            rv = normalize_str(r.get(region_col))
            if rv != '':
                rec['regions'].add(rv)
        if platform_col:
            pv = normalize_platform_value(r.get(platform_col))
            if pv != '':
                rec['platforms'].add(pv)
        if read_length_col:
            lv = normalize_str(r.get(read_length_col))
            if lv != '':
                rec['read_lengths'].add(lv)
        if sample_col:
            sv = normalize_str(r.get(sample_col))
            if sv != '':
                rec['samples'].add(sv)
    return subjects


def subset_meta_rows(
    meta_rows: List[Dict[str, str]],
    mapping: Dict[str, str],
    allowed_subjects: Optional[Set[str]] = None,
    allowed_samples: Optional[Set[str]] = None
) -> List[Dict[str, str]]:
    subject_col = mapping.get('subject_id', '')
    sample_col = mapping.get('sample_id', '')
    if not allowed_subjects and not allowed_samples:
        return meta_rows

    out = []
    for r in meta_rows:
        keep = False
        if allowed_subjects and subject_col:
            sid = normalize_str(r.get(subject_col))
            if sid in allowed_subjects:
                keep = True
        if allowed_samples and sample_col:
            samp = normalize_str(r.get(sample_col))
            if samp in allowed_samples:
                keep = True
        if keep:
            out.append(r)
    return out


def cohort_context(
    dataset: str,
    meta_rows: List[Dict[str, str]],
    mapping: Dict[str, str],
    allowed_subjects: Optional[Set[str]] = None,
    allowed_samples: Optional[Set[str]] = None,
    allowed_dx: Optional[Set[str]] = None
) -> Dict[str, str]:
    rows = subset_meta_rows(meta_rows, mapping, allowed_subjects=allowed_subjects, allowed_samples=allowed_samples)
    subjects = build_subject_records(rows, mapping)
    if allowed_dx:
        subjects = {k: v for k, v in subjects.items() if normalize_str(v.get('diagnosis')) in allowed_dx}

    age_vals = [v.get('age', '') for v in subjects.values()]
    sex_vals = [v.get('sex', '') for v in subjects.values()]
    region_vals = []
    platform_vals = []
    for v in subjects.values():
        region_vals.extend(sorted(v.get('regions', set())))
        platform_vals.extend(sorted(v.get('platforms', set())))

    platform_manual = mapping.get('platform_manual', '')
    platform_summary = summarize_unique(platform_vals)
    if dataset == 'Gandal 2022 bulk cortex':
        read_lengths = sorted({rl for v in subjects.values() for rl in v.get('read_lengths', set()) if rl != ''})
        if platform_summary == '' and platform_manual != '':
            platform_summary = platform_manual
        if read_lengths:
            read_len_txt = ', '.join(read_lengths)
            platform_summary = f'{platform_summary}; Read_Length={read_len_txt}' if platform_summary else f'Read_Length={read_len_txt}'
    elif platform_summary == '' and platform_manual != '':
        platform_summary = platform_manual

    return {
        'age_range_years': summarize_age_range(age_vals),
        'sex_distribution': summarize_distribution(sex_vals),
        'sampled_brain_regions': summarize_unique(region_vals),
        'sequencing_platform': platform_summary,
    }


def finalize_s1_row(
    row: Dict[str, str],
    dataset: str,
    meta_rows: List[Dict[str, str]],
    mapping: Dict[str, str],
    allowed_subjects: Optional[Set[str]] = None,
    allowed_samples: Optional[Set[str]] = None,
    allowed_dx: Optional[Set[str]] = None
) -> Dict[str, str]:
    rr = dict(row)
    rr.update(cohort_context(dataset, meta_rows, mapping, allowed_subjects=allowed_subjects, allowed_samples=allowed_samples, allowed_dx=allowed_dx))
    return rr


def extract_top_markers(preview: str, n: int = 10) -> str:
    s = normalize_str(preview)
    if s == '':
        return ''
    toks = [x for x in re.split(r'[,;|\s]+', s) if x]
    return '; '.join(toks[:n])


# -------------------------------------------------
# Package paths
# -------------------------------------------------
pkg1 = ROOT / 'Package1_Figure1_DiscoveryAudit'
pkg1b = ROOT / 'Package1b_ClusterAnnotation_ContaminantAudit'
pkg2 = ROOT / 'Package2_Define_Cluster2'
pkg2b = ROOT / 'Package2b_ResidualBackgroundAudit'
pkg3 = ROOT / 'Package3_Discovery_DonorAwareDiseaseAssociation'
pkg4b = ROOT / 'Package4b_DirectionConsistencySummary_v2_manual'
pkg4v2 = ROOT / 'Package4_CrossCohortValidation_Velmeshev_v2'
pkg5 = ROOT / 'Package5_DirectionalConcordance_GlobalSummary'
pkg7 = ROOT / 'Package7_MinimalRobustnessAnalysis'
pkg8 = ROOT / 'Package8_ExternalValidation_Gandal2022'
pkg8b = ROOT / 'Package8b_Gandal2022_LOO_Sensitivity'

# -------------------------------------------------
# Load cohort metadata
# -------------------------------------------------
discovery_meta_rows = read_tsv(COHORT_META['Discovery']['file'])
validation_meta_rows = read_tsv(COHORT_META['Velmeshev validation']['file'])
gandal_meta_rows = read_rdata_frame(COHORT_META['Gandal 2022 bulk cortex']['file'], COHORT_META['Gandal 2022 bulk cortex']['object_name'])

# -------------------------------------------------
# S1 Cohort summary (metadata-expanded)
# -------------------------------------------------
s1 = []

exact_n_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '00_Figure1_exact_n.tsv',
        pkg1 / '00_Figure1_exact_n.tsv',
    ],
    glob_name='00_Figure1_exact_n.tsv'
)

discovery_donor_prop_rows = read_tsv(pkg1 / '06_allcluster_donor_proportions.tsv.gz')
discovery_audited_subjects = extract_id_set(discovery_donor_prop_rows, ['donor', 'subject', 'Subject', 'individual_ID', 'iid', 'sample'])

retained_donor_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '05_retained_cluster_donor_proportions.tsv',
        pkg1 / '05_retained_cluster_donor_proportions.tsv',
        pkg3 / '05_retained_cluster_donor_proportions.tsv',
    ],
    glob_name='*retained*donor*proportion*.tsv*'
)
discovery_retained_subjects = extract_id_set(retained_donor_rows, ['donor', 'subject', 'Subject', 'individual_ID', 'iid', 'sample'])
if not discovery_retained_subjects:
    discovery_retained_subjects = discovery_audited_subjects

if exact_n_rows:
    grouped = {}
    for r in exact_n_rows:
        panel = normalize_str(r.get('panel'))
        diagnosis = normalize_str(r.get('diagnosis'))
        if panel == 'A/B':
            scope = 'Discovery audited microglial object'
        elif panel == 'D/E':
            scope = 'Discovery retained-core object'
        else:
            continue
        grouped.setdefault(scope, {})
        grouped[scope][diagnosis] = {
            'n_cells': maybe_int(r.get('n_cells')),
            'n_donors': maybe_int(r.get('n_donors')),
        }

    for scope, diag_map in grouped.items():
        all_map = diag_map.get('All', {})
        asd_map = diag_map.get('ASD', {})
        ctl_map = diag_map.get('Control', {})
        subject_subset = discovery_retained_subjects if scope == 'Discovery retained-core object' else discovery_audited_subjects
        note = 'Counts derived from exact n summary; demographic/context fields derived from source metadata aggregated at the subject level.'
        if not subject_subset:
            note += ' Donor identifiers were not recoverable from auxiliary tables, so metadata fields reflect the underlying discovery resource.'
        s1.append(finalize_s1_row({
            'dataset': 'Discovery',
            'summary_scope': scope,
            'n_ASD_cases': asd_map.get('n_donors', ''),
            'n_Control_cases': ctl_map.get('n_donors', ''),
            'n_total_cases': all_map.get('n_donors', ''),
            'qc_retained_subjects': all_map.get('n_donors', ''),
            'qc_retained_nuclei': all_map.get('n_cells', ''),
            'qc_retained_samples': '',
            'notes': note,
        }, 'Discovery', discovery_meta_rows, COHORT_META['Discovery'], allowed_subjects=subject_subset or None, allowed_dx={'ASD', 'Control'}))
else:
    rows_disc = read_tsv(pkg1 / '06_allcluster_donor_proportions.tsv.gz')
    donor_cols = unique_values(rows_disc, ['donor', 'sample', 'Subject', 'subject', 'individual_ID', 'iid'])
    dx_cols = unique_values(rows_disc, ['dx', 'Diagnosis', 'diagnosis', 'group', 'Group', 'dx_std'])
    if donor_cols and dx_cols:
        seen = {}
        for r in rows_disc:
            donor = normalize_str(r.get(donor_cols[0]))
            dx = canonical_dx(r.get(dx_cols[0]))
            if donor != '' and dx != '':
                seen[donor] = dx
        n_asd = sum(v == 'ASD' for v in seen.values())
        n_ctl = sum(v == 'Control' for v in seen.values())
        s1.append(finalize_s1_row({
            'dataset': 'Discovery',
            'summary_scope': 'Discovery audited microglial object',
            'n_ASD_cases': n_asd,
            'n_Control_cases': n_ctl,
            'n_total_cases': len(seen),
            'qc_retained_subjects': len(seen),
            'qc_retained_nuclei': '',
            'qc_retained_samples': '',
            'notes': 'Fallback donor split derived from donor-proportion table; metadata fields derived from subject-level source metadata.',
        }, 'Discovery', discovery_meta_rows, COHORT_META['Discovery'], allowed_subjects=set(seen.keys()), allowed_dx={'ASD', 'Control'}))

# Validation summary
val_metric_rows = read_tsv(pkg4b / '03_direction_consistency_summary_pretty.tsv') or read_tsv(pkg4b / '04_direction_consistency_summary_manuscript.tsv')
val_counts = first_metric_counts(val_metric_rows)
shared_validation_rows = read_tsv(pkg4b / '00_shared_comparable_samples.tsv')
val_shared_subjects = extract_id_set(shared_validation_rows, ['subject', 'Subject', 'individual', 'sample', 'donor'])
val_shared_samples = extract_id_set(shared_validation_rows, ['sample', 'Sample', 'sample_id'])

if val_counts['n_ASD'] is not None or val_counts['n_Control'] is not None:
    note = 'Strict shared donor set used for prespecified score-based validation metrics.'
    if not val_shared_subjects and not val_shared_samples:
        note += ' Donor/sample identifiers were not recoverable from the shared-comparable table, so metadata fields reflect the underlying validation resource.'
    s1.append(finalize_s1_row({
        'dataset': 'Velmeshev validation',
        'summary_scope': 'Strict shared validation set',
        'n_ASD_cases': val_counts['n_ASD'] or '',
        'n_Control_cases': val_counts['n_Control'] or '',
        'n_total_cases': (val_counts['n_ASD'] or 0) + (val_counts['n_Control'] or 0),
        'qc_retained_subjects': (val_counts['n_ASD'] or 0) + (val_counts['n_Control'] or 0),
        'qc_retained_nuclei': '',
        'qc_retained_samples': '',
        'notes': note,
    }, 'Velmeshev validation', validation_meta_rows, COHORT_META['Velmeshev validation'], allowed_subjects=val_shared_subjects or None, allowed_samples=val_shared_samples or None, allowed_dx={'ASD', 'Control'}))

val_micro = read_tsv(pkg4v2 / '03_validation_microglia_dx_counts.tsv')
if val_micro and {'dx', 'N'} <= set(val_micro[0].keys()):
    n_asd_cells = sum(maybe_int(r['N']) or 0 for r in val_micro if canonical_dx(r.get('dx')) == 'ASD')
    n_ctl_cells = sum(maybe_int(r['N']) or 0 for r in val_micro if canonical_dx(r.get('dx')) == 'Control')
    note = 'Validation microglial-cell counts retained for descriptive context.'
    if not val_shared_subjects:
        note += ' Metadata fields reflect the underlying validation resource.'
    s1.append(finalize_s1_row({
        'dataset': 'Velmeshev validation',
        'summary_scope': 'Validation microglia retained for descriptive context',
        'n_ASD_cases': '',
        'n_Control_cases': '',
        'n_total_cases': '',
        'qc_retained_subjects': '',
        'qc_retained_nuclei': n_asd_cells + n_ctl_cells,
        'qc_retained_samples': '',
        'notes': note,
    }, 'Velmeshev validation', validation_meta_rows, COHORT_META['Velmeshev validation'], allowed_subjects=val_shared_subjects or None, allowed_dx={'ASD', 'Control'}))

# Gandal summary
rows_gandal_scores = read_tsv(pkg8 / '07_sample_level_program_scores.tsv')
rows_gandal_dx = read_tsv(pkg8 / '03_filtered_dx_counts.tsv')
gandal_filtered_samples = extract_id_set(rows_gandal_scores, ['Sample', 'sample', 'sample_id'])
gandal_filtered_subjects = extract_id_set(rows_gandal_scores, ['Subject', 'subject'])

if rows_gandal_dx and {'dx', 'N'} <= set(rows_gandal_dx[0].keys()):
    n_asd = sum(maybe_int(r['N']) or 0 for r in rows_gandal_dx if canonical_dx(r.get('dx')) == 'ASD')
    n_ctl = sum(maybe_int(r['N']) or 0 for r in rows_gandal_dx if canonical_dx(r.get('dx')) == 'Control')
    note = 'Filtered external bulk-cortex validation samples. The published resource included ASD, control, and Dup15q samples; only the ASD/control subset retained after diagnosis filtering was used for the present external validation analyses.'
    s1.append(finalize_s1_row({
        'dataset': 'Gandal 2022 bulk cortex',
        'summary_scope': 'Filtered external bulk-cortex cohort',
        'n_ASD_cases': n_asd,
        'n_Control_cases': n_ctl,
        'n_total_cases': n_asd + n_ctl,
        'qc_retained_subjects': len(gandal_filtered_subjects) if gandal_filtered_subjects else '',
        'qc_retained_nuclei': '',
        'qc_retained_samples': n_asd + n_ctl,
        'notes': note,
    }, 'Gandal 2022 bulk cortex', gandal_meta_rows, COHORT_META['Gandal 2022 bulk cortex'], allowed_subjects=gandal_filtered_subjects or None, allowed_samples=gandal_filtered_samples or None, allowed_dx={'ASD', 'Control'}))

s1 = clean_rows(s1)

# -------------------------------------------------
# S2 Discovery audit / exclusion rationale (reworked)
# -------------------------------------------------
marker_support = read_tsv(pkg1 / '10_marker_support.tsv')
cluster_priority = read_tsv(pkg1 / '11_cluster_prioritization.tsv')
top_candidate = read_tsv(pkg1 / '12_top_candidate.tsv')
cluster2_summary = read_tsv(pkg1 / '13_cluster2_summary.tsv')
panelC_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '03_panelC_underlying_values.tsv',
        pkg1 / '03_panelC_underlying_values.tsv',
    ],
    glob_name='03_panelC_underlying_values.tsv'
)
disc_cluster_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '01_discovery_clusters.tsv',
        pkg1 / '01_discovery_clusters.tsv',
    ],
    glob_name='01_discovery_clusters.tsv'
)
retained_cluster_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '02_retained_clusters_from_strict_object.tsv',
        pkg1 / '02_retained_clusters_from_strict_object.tsv',
    ],
    glob_name='02_retained_clusters_from_strict_object.tsv'
)

retained_clusters = set()
for r in retained_cluster_rows:
    c = normalize_str(r.get('retained_cluster') or r.get('cluster'))
    if c != '':
        retained_clusters.add(c)
if not retained_clusters:
    retained_clusters = {'0', '2'}

marker_idx = rows_to_index(marker_support, 'cluster')
priority_idx = rows_to_index(cluster_priority, 'cluster')
panelC_idx = rows_to_index(panelC_rows, 'cluster')

cluster_ids = set(marker_idx.keys()) | set(priority_idx.keys()) | set(panelC_idx.keys())
for r in disc_cluster_rows:
    c = normalize_str(r.get('cluster'))
    if c != '':
        cluster_ids.add(c)

def sort_key(x):
    try:
        return (0, int(x))
    except Exception:
        return (1, x)

s2 = []
for c in sorted(cluster_ids, key=sort_key):
    row = {'section': 'cluster_audit', 'cluster': c}
    if c in marker_idx:
        row.update(marker_idx[c])
    if c in priority_idx:
        for k, v in priority_idx[c].items():
            if k not in row:
                row[k] = v
    if c in panelC_idx:
        for k, v in panelC_idx[c].items():
            if k not in row:
                row[k] = v
    row['final_decision'] = 'retained' if c in retained_clusters else 'excluded'
    row['primary_reason'] = infer_primary_reason(row, retained_clusters)
    s2.append(row)

s2_extra = []
for r in s2:
    if normalize_str(r.get('section')) != 'cluster_audit':
        continue
    s2_extra.append({
        'section': 'cluster_marker_context',
        'cluster': r.get('cluster', ''),
        'final_decision': r.get('final_decision', ''),
        'primary_reason': r.get('primary_reason', ''),
        'top10_markers': extract_top_markers(r.get('top_marker_preview') or r.get('top50_preview') or ''),
        'n_homeostatic_markers_in_top50': r.get('n_homeostatic_markers_in_top50', ''),
        'n_activation_markers_in_top50': r.get('n_activation_markers_in_top50', ''),
        'microglia_total_score': r.get('microglia_total_score', ''),
        'top_nonmg_score': r.get('top_nonmg_score', ''),
    })

spp1_context_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '20_spp1_all_major_celltypes.tsv',
        pkg1 / '20_spp1_all_major_celltypes.tsv',
    ],
    glob_name='*spp1*celltype*.tsv*'
)
if spp1_context_rows:
    s2_extra.extend(add_section(spp1_context_rows, 'spp1_context_all_major_cell_types'))

all_celltype_prop_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '21_all_celltype_proportions.tsv',
        pkg1 / '21_all_celltype_proportions.tsv',
    ],
    glob_name='*celltype*proportion*.tsv*'
)
if all_celltype_prop_rows:
    s2_extra.extend(add_section(all_celltype_prop_rows, 'all_celltype_proportions_context'))

s2.extend(clean_rows(s2_extra))

if top_candidate:
    for r in add_section(top_candidate, 'top_candidate'):
        s2.append(r)
if cluster2_summary:
    for r in add_section(cluster2_summary, 'cluster2_summary'):
        s2.append(r)

# -------------------------------------------------
# S3-S6 unchanged except cleaning
# -------------------------------------------------
s3 = combine_rows([
    add_section(read_tsv(pkg2 / 'tables' / '22_cluster2_core_summary_up.tsv'), 'candidate_program'),
    add_section(read_tsv(pkg2 / 'tables' / '23_cluster0_core_summary_up.tsv'), 'reference_program'),
])

s4 = clean_rows(read_tsv(pkg3 / '13_pseudobulk_all_retained_ASD_vs_Control.tsv.gz'))

s5 = combine_rows([
    add_section(read_tsv(pkg3 / '15_pseudobulk_cluster0_ASD_vs_Control.tsv.gz'), 'cluster0'),
    add_section(read_tsv(pkg3 / '17_pseudobulk_cluster2_ASD_vs_Control.tsv.gz'), 'cluster2'),
])

s6 = combine_rows([
    add_section(read_tsv(pkg4b / '00_shared_comparable_samples.tsv'), 'shared_comparable_samples'),
    add_section(read_tsv(pkg4b / '00b_shared_sample_counts.tsv'), 'shared_sample_counts'),
    add_section(read_tsv(pkg4b / '03_direction_consistency_summary_pretty.tsv'), 'direction_consistency_pretty'),
    add_section(read_tsv(pkg4b / '04_direction_consistency_summary_manuscript.tsv'), 'direction_consistency_manuscript'),
])

# -------------------------------------------------
# S7-S9 flattened
# -------------------------------------------------
s7 = combine_rows([
    add_section(read_tsv(pkg5 / '01_direction_concordance_master.tsv'), 'direction_concordance_master'),
    add_section(read_tsv(pkg5 / '02_global_direction_tests.tsv'), 'global_direction_tests'),
    add_section(read_tsv(pkg5 / '03_metrics_ranked_by_nominal_p.tsv'), 'metrics_ranked_by_nominal_p'),
    add_section(read_tsv(pkg5 / '04_metrics_ranked_by_aligned_effect.tsv'), 'metrics_ranked_by_aligned_effect'),
    add_section(read_tsv(pkg7 / '01_score_defined_harmonization_sensitivity.tsv'), 'harmonization_sensitivity'),
    add_section(read_tsv(pkg7 / '02_score_defined_direction_stability.tsv'), 'direction_stability'),
    add_section(read_tsv(pkg7 / '03_leave_one_metric_out_global_summary.tsv'), 'leave_one_metric_out_summary'),
])

s8 = combine_rows([
    add_section(read_txt_as_rows(pkg8 / '04_candidate_genes_used.txt', 'gene', {'set_name': 'candidate_genes_used'}), 'gene_membership'),
    add_section(read_txt_as_rows(pkg8 / '04_reference_genes_used.txt', 'gene', {'set_name': 'reference_genes_used'}), 'gene_membership'),
    add_section(read_txt_as_rows(pkg8 / '06_candidate_hit_genes.txt', 'gene', {'set_name': 'candidate_hit_genes'}), 'gene_membership'),
    add_section(read_txt_as_rows(pkg8 / '06_reference_hit_genes.txt', 'gene', {'set_name': 'reference_hit_genes'}), 'gene_membership'),
    add_section(read_tsv(pkg8 / '06_gene_set_overlap.tsv'), 'gene_set_overlap'),
])

s9 = combine_rows([
    add_section(read_tsv(pkg8b / '04_leave_one_region_out_results.tsv'), 'leave_one_region_out_results'),
    add_section(read_tsv(pkg8b / '05_leave_one_region_out_vs_baseline.tsv'), 'leave_one_region_out_vs_baseline'),
    add_section(read_tsv(pkg8b / '06_leave_one_region_out_stability_summary.tsv'), 'leave_one_region_out_stability_summary'),
])

# -------------------------------------------------
# S10 Structured software / package / parameter summary
# -------------------------------------------------
session_files = [
    pkg1 / 'sessionInfo_package1.txt',
    pkg1b / 'sessionInfo_package1b.txt',
    pkg2 / 'sessionInfo_package2.txt',
    pkg2b / 'sessionInfo_package2b.txt',
    pkg4v2 / '31_sessionInfo.txt',
    pkg8 / '13_sessionInfo.txt',
    pkg8b / '11_sessionInfo.txt',
]
session_files = [p for p in session_files if p.exists()]

env_source = session_files[0] if session_files else None
env_info = parse_session_environment(env_source) if env_source else {}
pkg_versions = parse_package_versions(session_files)

s10 = []

# Environment
for item in ['R', 'Platform', 'Operating system']:
    if item in env_info:
        s10.append({
            'Category': 'Software environment',
            'analysis_stage': 'Global',
            'Item': item,
            'Version_or_value': env_info[item],
            'Purpose_or_definition': 'Core runtime environment',
            'Notes': ''
        })

# Packages
for pkg in CURATED_PACKAGES:
    ver = pkg_versions.get(pkg, '')
    if ver == '':
        continue
    s10.append({
        'Category': 'Package version',
        'analysis_stage': 'Global',
        'Item': pkg,
        'Version_or_value': ver,
        'Purpose_or_definition': PACKAGE_PURPOSE.get(pkg, ''),
        'Notes': ''
    })

# Discovery parameters
disc_all = next((r for r in exact_n_rows if normalize_str(r.get('panel')) == 'A/B' and normalize_str(r.get('diagnosis')) == 'All'), None)
disc_retained = next((r for r in exact_n_rows if normalize_str(r.get('panel')) == 'D/E' and normalize_str(r.get('diagnosis')) == 'All'), None)

disc_cluster_count = None
if disc_all:
    disc_cluster_count = maybe_int(disc_all.get('n_clusters'))
elif disc_cluster_rows:
    disc_cluster_count = len(disc_cluster_rows)

retained_cluster_count = None
if disc_retained:
    retained_cluster_count = maybe_int(disc_retained.get('n_clusters'))
elif retained_cluster_rows:
    retained_cluster_count = len(retained_cluster_rows)

s10.extend([
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Discovery clustering resolution',
        'Version_or_value': MANUAL_S10_VALUES['discovery_clustering_resolution'],
        'Purpose_or_definition': 'Resolution used to define the reclustered discovery microglial analysis space.',
        'Notes': MANUAL_S10_VALUES['discovery_reclustering_rationale'],
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Discovery clusters',
        'Version_or_value': disc_cluster_count if disc_cluster_count is not None else '',
        'Purpose_or_definition': 'Number of clusters in the audited discovery-stage microglial analysis space.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Retained clusters',
        'Version_or_value': ', '.join(sorted(retained_clusters, key=sort_key)),
        'Purpose_or_definition': 'Clusters retained as the cortical microglial core for downstream inference.',
        'Notes': '',
    },
])

if disc_all:
    s10.append({
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Discovery audited microglial object',
        'Version_or_value': f"{disc_all.get('n_cells')} cells from {disc_all.get('n_donors')} donors",
        'Purpose_or_definition': 'Audited discovery-stage microglial analysis object.',
        'Notes': '',
    })
if disc_retained:
    s10.append({
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Strict retained-core object',
        'Version_or_value': f"{disc_retained.get('n_cells')} cells from {disc_retained.get('n_donors')} donors",
        'Purpose_or_definition': 'Strict retained-core microglial object used for downstream analyses.',
        'Notes': '',
    })

# Validation parameters
s10.append({
    'Category': 'Analysis parameter',
    'analysis_stage': 'Validation',
    'Item': 'Validation reclustering resolution',
    'Version_or_value': MANUAL_S10_VALUES['validation_reclustering_resolution'],
    'Purpose_or_definition': 'Resolution used for validation microglia reclustering in the Velmeshev dataset.',
    'Notes': MANUAL_S10_VALUES['validation_reclustering_rationale'],
})

s10.extend([
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Validation',
        'Item': 'Validation PCs used',
        'Version_or_value': MANUAL_S10_VALUES['validation_npcs'],
        'Purpose_or_definition': 'Number of principal components used for validation microglia reclustering.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Validation',
        'Item': 'Score-defined candidate quantile',
        'Version_or_value': MANUAL_S10_VALUES['candidate_quantile'],
        'Purpose_or_definition': 'Quantile threshold used to define score-based candidate-like cells in validation analyses.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Validation',
        'Item': 'Score-defined reference quantile',
        'Version_or_value': MANUAL_S10_VALUES['reference_quantile'],
        'Purpose_or_definition': 'Quantile threshold used to define score-based reference-like cells in validation analyses.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Validation',
        'Item': 'Mapped-cluster correlation-delta threshold',
        'Version_or_value': MANUAL_S10_VALUES['mapping_corr_delta_threshold'],
        'Purpose_or_definition': 'Threshold used to summarize weak mapped-cluster preference toward discovery cluster 2.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Validation',
        'Item': 'Mapped-cluster state-delta threshold',
        'Version_or_value': MANUAL_S10_VALUES['mapping_state_delta_threshold'],
        'Purpose_or_definition': 'Threshold used to summarize mapped-cluster state preference in validation sensitivity summaries.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Validation',
        'Item': 'Minimum cells per sample',
        'Version_or_value': MANUAL_S10_VALUES['min_cells_sample'],
        'Purpose_or_definition': 'Minimum cells required per donor/sample for validation score-based aggregation.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Validation',
        'Item': 'Minimum cells per score-defined group',
        'Version_or_value': MANUAL_S10_VALUES['min_cells_group'],
        'Purpose_or_definition': 'Minimum cells required per score-defined candidate or reference subset.',
        'Notes': '',
    },
])

if val_counts['n_ASD'] is not None or val_counts['n_Control'] is not None:
    s10.append({
        'Category': 'Analysis parameter',
        'analysis_stage': 'Validation',
        'Item': 'Strict shared validation set',
        'Version_or_value': f"ASD={val_counts['n_ASD']}; Control={val_counts['n_Control']}",
        'Purpose_or_definition': 'Shared donor set used for the seven prespecified score-based validation metrics.',
        'Notes': '',
    })

global_dir_tests = read_tsv(pkg5 / '02_global_direction_tests.tsv')
n_metrics = ''
for r in global_dir_tests:
    if normalize_str(r.get('test_name')) == 'n_total_metrics':
        n_metrics = r.get('value', '')
        break
s10.append({
    'Category': 'Analysis parameter',
    'analysis_stage': 'Validation',
    'Item': 'Prespecified score-based validation metrics',
    'Version_or_value': n_metrics,
    'Purpose_or_definition': 'Number of primary score-based validation readouts used for directional-concordance testing.',
    'Notes': '',
})

# Bulk parameters
s10.extend([
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'External bulk cortex',
        'Item': 'External bulk validation dataset',
        'Version_or_value': 'Gandal 2022 bulk cortex',
        'Purpose_or_definition': 'Independent tissue-level validation cohort.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'External bulk cortex',
        'Item': 'Primary external scores',
        'Version_or_value': 'candidate score; candidate-minus-reference score; reference score',
        'Purpose_or_definition': 'Bulk projection readouts derived from the frozen candidate/reference programs.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'External bulk cortex',
        'Item': 'Bulk model covariates',
        'Version_or_value': MANUAL_S10_VALUES['bulk_model_covariates'],
        'Purpose_or_definition': 'Covariates included in the mixed-model ASD effect testing framework.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'External bulk cortex',
        'Item': 'Sensitivity analysis',
        'Version_or_value': 'leave-one-region-out',
        'Purpose_or_definition': 'Region-wise robustness analysis for the external bulk-cortex validation.',
        'Notes': '',
    },
])

# Reproducibility
s10.extend([
    {
        'Category': 'Reproducibility resource',
        'analysis_stage': 'Global',
        'Item': 'GitHub repository',
        'Version_or_value': MANUAL_S10_VALUES['github_repository'],
        'Purpose_or_definition': 'Analysis scripts and figure-generation code.',
        'Notes': '',
    },
    {
        'Category': 'Reproducibility resource',
        'analysis_stage': 'Global',
        'Item': 'Zenodo archive',
        'Version_or_value': MANUAL_S10_VALUES['zenodo_archive'],
        'Purpose_or_definition': 'Versioned archival release linked to the submitted manuscript.',
        'Notes': 'Replace with the final DOI or persistent archive URL before submission if still pending.',
    },
])

# -------------------------------------------------
# Write workbook
# -------------------------------------------------
wb = Workbook()
wb.remove(wb.active)

sheet_map = [
    ('S1_CohortSummary', s1),
    ('S2_DiscoveryAudit', s2),
    ('S3_GenePrograms', s3),
    ('S4_PooledDEGs', s4),
    ('S5_ClusterDEGs', s5),
    ('S6_ValidationMetrics', s6),
    ('S7_ConcordanceRobust', s7),
    ('S8_GandalMapping', s8),
    ('S9_LOO_Results', s9),
    ('S10_SoftwareParams', s10),
]

for name, rows in sheet_map:
    ws = wb.create_sheet(title=name[:31])
    write_sheet(ws, rows)

wb.save(OUTFILE)
print(f'Wrote workbook: {OUTFILE}')
