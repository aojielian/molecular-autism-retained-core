from pathlib import Path
from copy import copy
import os
import csv
import gzip
import json
import tempfile
import subprocess
from collections import Counter
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================
# User-adjustable paths
# =========================
ROOT = Path(os.environ.get('MA_REVISION_ROOT', '/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/MolecularAutism_Revision'))
INPUT_XLSX = Path(os.environ.get('MA_INPUT_XLSX', 'Supplementary_Tables_revised.xlsx'))
OUTPUT_XLSX = Path(os.environ.get('MA_OUTPUT_XLSX', str(INPUT_XLSX.with_name('Supplementary_Tables_revised_S1fixed.xlsx'))))

DISCOVERY_META = Path('/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/PsychENCODE_Science_2024/meta.tsv')
VEL_META = Path('/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/Velmeshev_scRNA/meta.tsv')
GANDAL_RDATA = Path('/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/Gandal_2022/Gene_NormalizedExpression_Metadata_wModelMatrix.RData')

PKG1 = ROOT / 'Package1_Figure1_DiscoveryAudit'
PKG4B = ROOT / 'Package4b_DirectionConsistencySummary_v2_manual'
PKG4V2 = ROOT / 'Package4_CrossCohortValidation_Velmeshev_v2'
PKG8 = ROOT / 'Package8_ExternalValidation_Gandal2022'

STRICT_SHARED_ASD = int(os.environ.get('MA_STRICT_SHARED_ASD', '11'))
STRICT_SHARED_CONTROL = int(os.environ.get('MA_STRICT_SHARED_CONTROL', '14'))

TARGET_HEADERS = [
    'dataset', 'summary_scope', 'n_ASD_cases', 'n_Control_cases', 'n_total_cases',
    'qc_retained_subjects', 'qc_retained_nuclei', 'qc_retained_samples',
    'age_range_years', 'sex_distribution', 'sampled_brain_regions',
    'sequencing_platform', 'notes'
]

TARGET_ORDER = [
    ('Discovery', 'Discovery audited microglial object'),
    ('Discovery', 'Discovery retained-core object'),
    ('Velmeshev validation', 'Strict shared validation set'),
    ('Velmeshev validation', 'Validation microglia retained for descriptive context'),
    ('Gandal 2022 bulk cortex', 'Filtered external bulk-cortex cohort'),
]


# =========================
# IO helpers
# =========================
def read_tsv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    opener = gzip.open if path.suffix == '.gz' else open
    with opener(path, 'rt', encoding='utf-8', newline='') as f:
        return [dict(r) for r in csv.DictReader(f, delimiter='\t')]


def read_rdata_datmeta(path: Path, object_name: str = 'datMeta') -> List[Dict[str, str]]:
    if not path.exists():
        return []
    tmp = tempfile.NamedTemporaryFile(prefix='datMeta_', suffix='.tsv', delete=False)
    tmp_path = Path(tmp.name)
    tmp.close()
    r_code = f"""
    e <- new.env()
    load({json.dumps(str(path))}, envir = e)
    obj <- get({json.dumps(object_name)}, envir = e)
    if (!is.data.frame(obj)) obj <- as.data.frame(obj, stringsAsFactors = FALSE)
    write.table(obj, file = {json.dumps(str(tmp_path))}, sep = "\\t", quote = FALSE, row.names = FALSE, col.names = TRUE, na = "")
    """
    try:
        subprocess.run(['Rscript', '-e', r_code], check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        rows = read_tsv(tmp_path)
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
    return rows


def maybe_int(x):
    try:
        if x is None or str(x).strip() == '':
            return None
        return int(float(str(x)))
    except Exception:
        return None


def maybe_float(x):
    try:
        if x is None or str(x).strip() == '':
            return None
        return float(str(x))
    except Exception:
        return None


def norm(x) -> str:
    return '' if x is None else str(x).strip()


# =========================
# Domain helpers
# =========================
def canonical_dx(x: str) -> str:
    s = norm(x).lower()
    if s in {'asd', 'autism', 'autism spectrum disorder'}:
        return 'ASD'
    if s in {'ctl', 'control', 'ctrl', 'no known disorder'}:
        return 'Control'
    return ''


def sex_label(x: str) -> str:
    s = norm(x)
    if s == 'XX':
        return 'F'
    if s == 'XY':
        return 'M'
    if s in {'F', 'M'}:
        return s
    if s == 'XYY':
        return 'XYY'
    return s


def summarize_sex(counter: Counter) -> str:
    f = counter.get('F', 0)
    m = counter.get('M', 0)
    extras = []
    if counter.get('XYY', 0):
        extras.append(f"one donor with XYY karyotype")
    for k, v in sorted(counter.items()):
        if k in {'F', 'M', 'XYY', ''}:
            continue
        extras.append(f"{k}={v}")
    base = [f"F={f}", f"M={m}"]
    return '; '.join(base + extras)


def summarize_ages(values: List[float]) -> str:
    vals = sorted(v for v in values if v is not None)
    if not vals:
        return ''
    lo, hi = vals[0], vals[-1]
    def fmt(v):
        if abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return ('%.1f' % v).rstrip('0').rstrip('.')
    return f"{fmt(lo)}-{fmt(hi)}"


def summarize_unique(values: List[str]) -> str:
    vals = sorted({norm(v) for v in values if norm(v)})
    return '; '.join(vals)


def summarize_platform(values: List[str], dataset: str) -> str:
    vals = [norm(v) for v in values if norm(v)]
    uniq = sorted(set(vals))
    if dataset == 'Discovery':
        mapped = []
        for v in uniq:
            if v.upper() == 'V2':
                mapped.append('10x Genomics Chromium V2')
            elif v.upper() == 'V3':
                mapped.append('10x Genomics Chromium V3')
            else:
                mapped.append(v)
        return '; '.join(mapped)
    if dataset == 'Velmeshev validation':
        return 'single-nucleus RNA-seq'
    if dataset == 'Gandal 2022 bulk cortex':
        if not uniq:
            return 'bulk RNA-seq'
        return '; '.join(sorted(set(('bulk RNA-seq (' + v + ')') if 'bulk' not in v.lower() else v for v in uniq)))
    return '; '.join(uniq)


def existing_row_map(ws) -> Dict[Tuple[str, str], Dict[str, str]]:
    header = [norm(c.value) for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = {header[i]: ('' if row[i] is None else str(row[i])) for i in range(min(len(header), len(row)))}
        key = (vals.get('dataset', ''), vals.get('summary_scope', ''))
        if key != ('', ''):
            out[key] = vals
    return out


def find_first_existing(paths: List[Path]) -> Optional[Path]:
    for p in paths:
        if p.exists():
            return p
    return None


# =========================
# Metadata summarizers
# =========================
def discovery_subject_demo(meta_rows: List[Dict[str, str]], allowed_subjects: Optional[set] = None) -> Dict[str, str]:
    by_subject = {}
    for r in meta_rows:
        sid = norm(r.get('individual_ID'))
        if not sid:
            continue
        if allowed_subjects is not None and sid not in allowed_subjects:
            continue
        dx = canonical_dx(r.get('Diagnosis'))
        if dx not in {'ASD', 'Control'}:
            continue
        by_subject.setdefault(sid, {
            'Age': maybe_float(r.get('Age')),
            'Sex': sex_label(r.get('Sex_Chromosome')),
            'Brain_Region': set(),
            'Platform': set(),
        })
        if by_subject[sid]['Age'] is None:
            by_subject[sid]['Age'] = maybe_float(r.get('Age'))
        sx = sex_label(r.get('Sex_Chromosome'))
        if sx:
            by_subject[sid]['Sex'] = sx
        br = norm(r.get('Brain_Region'))
        if br:
            by_subject[sid]['Brain_Region'].add(br)
        pf = norm(r.get('Chemistry10X'))
        if pf:
            by_subject[sid]['Platform'].add(pf)
    ages = [v['Age'] for v in by_subject.values() if v['Age'] is not None]
    sex_ct = Counter(v['Sex'] for v in by_subject.values() if v['Sex'])
    regions = []
    platforms = []
    for v in by_subject.values():
        regions.extend(list(v['Brain_Region']))
        platforms.extend(list(v['Platform']))
    return {
        'age_range_years': summarize_ages(ages),
        'sex_distribution': summarize_sex(sex_ct),
        'sampled_brain_regions': summarize_unique(regions),
        'sequencing_platform': summarize_platform(platforms, 'Discovery'),
    }


def vel_subject_demo(meta_rows: List[Dict[str, str]], allowed_subjects: Optional[set] = None, allowed_samples: Optional[set] = None) -> Dict[str, str]:
    by_subject = {}
    sample_set = set()
    for r in meta_rows:
        sid = norm(r.get('individual'))
        sample = norm(r.get('sample'))
        if allowed_subjects is not None and sid not in allowed_subjects:
            continue
        if allowed_samples is not None and sample not in allowed_samples:
            continue
        dx = canonical_dx(r.get('diagnosis'))
        if dx not in {'ASD', 'Control'}:
            continue
        sample_set.add(sample)
        by_subject.setdefault(sid, {
            'Age': maybe_float(r.get('age')),
            'Sex': sex_label(r.get('sex')),
            'Region': set(),
        })
        br = norm(r.get('region'))
        if br:
            by_subject[sid]['Region'].add(br)
    ages = [v['Age'] for v in by_subject.values() if v['Age'] is not None]
    sex_ct = Counter(v['Sex'] for v in by_subject.values() if v['Sex'])
    regions = []
    for v in by_subject.values():
        regions.extend(list(v['Region']))
    return {
        'age_range_years': summarize_ages(ages),
        'sex_distribution': summarize_sex(sex_ct),
        'sampled_brain_regions': summarize_unique(regions),
        'sequencing_platform': 'single-nucleus RNA-seq',
        'subject_count': len(by_subject),
        'sample_count': len(sample_set),
    }


def gandal_subject_demo(meta_rows: List[Dict[str, str]], allowed_samples: Optional[set] = None) -> Dict[str, str]:
    by_subject = {}
    sample_set = set()
    platform_vals = []
    read_lengths = []
    for r in meta_rows:
        sample = norm(r.get('sample_id'))
        if allowed_samples is not None and sample not in allowed_samples:
            continue
        dx = canonical_dx(r.get('Diagnosis'))
        if dx not in {'ASD', 'Control'}:
            continue
        subj = norm(r.get('subject'))
        sample_set.add(sample)
        by_subject.setdefault(subj, {
            'Age': maybe_float(r.get('Age')),
            'Sex': sex_label(r.get('Sex')),
            'Region': set(),
        })
        br = norm(r.get('region'))
        if br:
            by_subject[subj]['Region'].add(br)
        if norm(r.get('SeqMethod')):
            platform_vals.append(norm(r.get('SeqMethod')))
        if norm(r.get('Read_Length')):
            read_lengths.append(norm(r.get('Read_Length')))
    ages = [v['Age'] for v in by_subject.values() if v['Age'] is not None]
    sex_ct = Counter(v['Sex'] for v in by_subject.values() if v['Sex'])
    regions = []
    for v in by_subject.values():
        regions.extend(list(v['Region']))
    pf = summarize_platform(platform_vals, 'Gandal 2022 bulk cortex')
    rl = sorted(set(read_lengths))
    if rl:
        pf = f"{pf}; Read_Length={', '.join(rl)}"
    return {
        'age_range_years': summarize_ages(ages),
        'sex_distribution': summarize_sex(sex_ct),
        'sampled_brain_regions': summarize_unique(regions),
        'sequencing_platform': pf,
        'subject_count': len(by_subject),
        'sample_count': len(sample_set),
    }


# =========================
# Data derivation
# =========================
def load_discovery_rows(existing: Dict[Tuple[str, str], Dict[str, str]], discovery_meta_rows: List[Dict[str, str]]) -> Tuple[Dict[str, str], Dict[str, str]]:
    aud_key = ('Discovery', 'Discovery audited microglial object')
    ret_key = ('Discovery', 'Discovery retained-core object')
    aud = dict(existing.get(aud_key, {}))
    ret = dict(existing.get(ret_key, {}))

    audited_demo = discovery_subject_demo(discovery_meta_rows)
    retained_demo = discovery_subject_demo(discovery_meta_rows)  # subject subset not reliably available here; keep metadata context consistent

    aud_row = {
        'dataset': 'Discovery',
        'summary_scope': 'Discovery audited microglial object',
        'n_ASD_cases': aud.get('n_ASD_cases', '31'),
        'n_Control_cases': aud.get('n_Control_cases', '29'),
        'n_total_cases': aud.get('n_total_cases', '60'),
        'qc_retained_subjects': aud.get('qc_retained_subjects', aud.get('n_total_cases', '60')),
        'qc_retained_nuclei': aud.get('qc_retained_nuclei', '34945'),
        'qc_retained_samples': aud.get('qc_retained_samples', ''),
        'age_range_years': aud.get('age_range_years', audited_demo['age_range_years']),
        'sex_distribution': aud.get('sex_distribution', audited_demo['sex_distribution']),
        'sampled_brain_regions': aud.get('sampled_brain_regions', audited_demo['sampled_brain_regions']),
        'sequencing_platform': aud.get('sequencing_platform', audited_demo['sequencing_platform']),
        'notes': aud.get('notes', 'Counts derived from exact n summary; demographic/context fields derived from source metadata aggregated at the subject level.'),
    }
    ret_row = {
        'dataset': 'Discovery',
        'summary_scope': 'Discovery retained-core object',
        'n_ASD_cases': ret.get('n_ASD_cases', '31'),
        'n_Control_cases': ret.get('n_Control_cases', '28'),
        'n_total_cases': ret.get('n_total_cases', '59'),
        'qc_retained_subjects': ret.get('qc_retained_subjects', ret.get('n_total_cases', '59')),
        'qc_retained_nuclei': ret.get('qc_retained_nuclei', '13985'),
        'qc_retained_samples': ret.get('qc_retained_samples', ''),
        'age_range_years': ret.get('age_range_years', retained_demo['age_range_years']),
        'sex_distribution': ret.get('sex_distribution', retained_demo['sex_distribution']),
        'sampled_brain_regions': ret.get('sampled_brain_regions', retained_demo['sampled_brain_regions']),
        'sequencing_platform': ret.get('sequencing_platform', retained_demo['sequencing_platform']),
        'notes': ret.get('notes', 'Counts derived from exact n summary; demographic/context fields derived from source metadata aggregated at the subject level.'),
    }
    return aud_row, ret_row


def load_validation_rows(existing: Dict[Tuple[str, str], Dict[str, str]], vel_meta_rows: List[Dict[str, str]]) -> Tuple[Dict[str, str], Dict[str, str]]:
    strict_key = ('Velmeshev validation', 'Strict shared validation set')
    desc_key = ('Velmeshev validation', 'Validation microglia retained for descriptive context')
    strict_old = dict(existing.get(strict_key, {}))
    desc_old = dict(existing.get(desc_key, {}))

    # Descriptive context from full validation metadata
    full_demo = vel_subject_demo(vel_meta_rows)
    full_asd = 0
    full_ctl = 0
    full_subject_dx = {}
    for r in vel_meta_rows:
        sid = norm(r.get('individual'))
        dx = canonical_dx(r.get('diagnosis'))
        if sid and dx in {'ASD', 'Control'}:
            full_subject_dx[sid] = dx
    full_asd = sum(v == 'ASD' for v in full_subject_dx.values())
    full_ctl = sum(v == 'Control' for v in full_subject_dx.values())

    # Strict shared sample counts if file exists; otherwise leave sample count blank
    shared_file = find_first_existing([
        PKG4B / '00_shared_comparable_samples.tsv',
        PKG4B / '00_shared_comparable_samples.tsv.gz',
    ])
    strict_sample_count = ''
    if shared_file is not None:
        rows = read_tsv(shared_file)
        subj = set()
        samp = set()
        for r in rows:
            for c in ['subject', 'Subject', 'individual', 'donor']:
                if norm(r.get(c)):
                    subj.add(norm(r.get(c)))
                    break
            for c in ['sample', 'Sample', 'sample_id']:
                if norm(r.get(c)):
                    samp.add(norm(r.get(c)))
                    break
        if samp:
            strict_sample_count = str(len(samp))

    # Microglial nuclei retained for descriptive context
    nuclei_file = find_first_existing([
        PKG4V2 / '03_validation_microglia_dx_counts.tsv',
        PKG4V2 / '03_validation_microglia_dx_counts.tsv.gz',
    ])
    total_nuclei = ''
    if nuclei_file is not None:
        rows = read_tsv(nuclei_file)
        total = 0
        for r in rows:
            dx = canonical_dx(r.get('dx'))
            if dx in {'ASD', 'Control'}:
                total += maybe_int(r.get('N')) or 0
        total_nuclei = str(total)

    strict_row = {
        'dataset': 'Velmeshev validation',
        'summary_scope': 'Strict shared validation set',
        'n_ASD_cases': str(STRICT_SHARED_ASD),
        'n_Control_cases': str(STRICT_SHARED_CONTROL),
        'n_total_cases': str(STRICT_SHARED_ASD + STRICT_SHARED_CONTROL),
        'qc_retained_subjects': str(STRICT_SHARED_ASD + STRICT_SHARED_CONTROL),
        'qc_retained_nuclei': '',
        'qc_retained_samples': strict_sample_count,
        'age_range_years': strict_old.get('age_range_years', full_demo['age_range_years']),
        'sex_distribution': strict_old.get('sex_distribution', full_demo['sex_distribution']),
        'sampled_brain_regions': strict_old.get('sampled_brain_regions', full_demo['sampled_brain_regions']),
        'sequencing_platform': strict_old.get('sequencing_platform', full_demo['sequencing_platform']),
        'notes': 'Strict shared donor set used for prespecified score-based validation metrics; counts correspond to the analytic intersection across all prespecified score-based validation metrics.',
    }
    desc_row = {
        'dataset': 'Velmeshev validation',
        'summary_scope': 'Validation microglia retained for descriptive context',
        'n_ASD_cases': str(full_asd),
        'n_Control_cases': str(full_ctl),
        'n_total_cases': str(full_asd + full_ctl),
        'qc_retained_subjects': str(full_demo['subject_count']),
        'qc_retained_nuclei': total_nuclei,
        'qc_retained_samples': str(full_demo['sample_count']),
        'age_range_years': desc_old.get('age_range_years', full_demo['age_range_years']),
        'sex_distribution': desc_old.get('sex_distribution', full_demo['sex_distribution']),
        'sampled_brain_regions': desc_old.get('sampled_brain_regions', full_demo['sampled_brain_regions']),
        'sequencing_platform': desc_old.get('sequencing_platform', full_demo['sequencing_platform']),
        'notes': 'Underlying validation resource used for descriptive context; this row is not the strict shared analytic set.',
    }
    return strict_row, desc_row


def load_gandal_row(existing: Dict[Tuple[str, str], Dict[str, str]], gandal_meta_rows: List[Dict[str, str]]) -> Dict[str, str]:
    key = ('Gandal 2022 bulk cortex', 'Filtered external bulk-cortex cohort')
    old = dict(existing.get(key, {}))

    filtered_score_file = find_first_existing([
        PKG8 / '07_sample_level_program_scores.tsv',
        PKG8 / '07_sample_level_program_scores.tsv.gz',
    ])
    filtered_dx_file = find_first_existing([
        PKG8 / '03_filtered_dx_counts.tsv',
        PKG8 / '03_filtered_dx_counts.tsv.gz',
    ])

    allowed_samples = None
    if filtered_score_file is not None:
        rows = read_tsv(filtered_score_file)
        sample_set = set()
        for r in rows:
            for c in ['Sample', 'sample', 'sample_id']:
                if norm(r.get(c)):
                    sample_set.add(norm(r.get(c)))
                    break
        if sample_set:
            allowed_samples = sample_set

    demo = gandal_subject_demo(gandal_meta_rows, allowed_samples=allowed_samples)

    n_asd, n_ctl = old.get('n_ASD_cases', '49'), old.get('n_Control_cases', '54')
    if filtered_dx_file is not None:
        rows = read_tsv(filtered_dx_file)
        asd = sum((maybe_int(r.get('N')) or 0) for r in rows if canonical_dx(r.get('dx')) == 'ASD')
        ctl = sum((maybe_int(r.get('N')) or 0) for r in rows if canonical_dx(r.get('dx')) == 'Control')
        if asd or ctl:
            n_asd, n_ctl = str(asd), str(ctl)

    total = ''
    if n_asd != '' and n_ctl != '':
        total = str((maybe_int(n_asd) or 0) + (maybe_int(n_ctl) or 0))

    return {
        'dataset': 'Gandal 2022 bulk cortex',
        'summary_scope': 'Filtered external bulk-cortex cohort',
        'n_ASD_cases': n_asd,
        'n_Control_cases': n_ctl,
        'n_total_cases': total,
        'qc_retained_subjects': old.get('qc_retained_subjects', str(demo['subject_count'])),
        'qc_retained_nuclei': '',
        'qc_retained_samples': old.get('qc_retained_samples', str(demo['sample_count'])),
        'age_range_years': old.get('age_range_years', demo['age_range_years']),
        'sex_distribution': old.get('sex_distribution', demo['sex_distribution']),
        'sampled_brain_regions': old.get('sampled_brain_regions', demo['sampled_brain_regions']),
        'sequencing_platform': old.get('sequencing_platform', demo['sequencing_platform']),
        'notes': 'Filtered external bulk-cortex validation samples. The published resource included ASD, control, and Dup15q samples; only the ASD/control subset retained after diagnosis filtering was used for the present external validation analyses.',
    }


# =========================
# Workbook writing
# =========================
def apply_header_style(ws):
    fill = PatternFill('solid', fgColor='FF1F4E78')
    font = Font(color='FFFFFFFF', bold=True)
    border = Border(
        left=Side(style='thin', color='FFD9E1F2'),
        right=Side(style='thin', color='FFD9E1F2'),
        top=Side(style='thin', color='FFD9E1F2'),
        bottom=Side(style='thin', color='FFD9E1F2'),
    )
    for col_idx, header in enumerate(TARGET_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border


def set_widths(ws):
    widths = {
        'A': 24, 'B': 42, 'C': 13, 'D': 16, 'E': 13, 'F': 20, 'G': 18,
        'H': 18, 'I': 16, 'J': 34, 'K': 46, 'L': 42, 'M': 86,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def remove_existing_tables(ws):
    if ws.tables:
        for name in list(ws.tables.keys()):
            del ws.tables[name]


def rewrite_s1_sheet(ws, rows: List[Dict[str, str]]):
    ws.delete_rows(1, ws.max_row)
    apply_header_style(ws)
    for i, row in enumerate(rows, start=2):
        for j, header in enumerate(TARGET_HEADERS, start=1):
            val = row.get(header, '')
            ws.cell(row=i, column=j, value=None if val == '' else val)
    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False
    set_widths(ws)
    for r in range(2, 2 + len(rows)):
        for c in range(1, len(TARGET_HEADERS) + 1):
            ws.cell(r, c).alignment = Alignment(vertical='top', wrap_text=True)
    remove_existing_tables(ws)
    table = Table(displayName='S1_CohortSummary_Table', ref=f"A1:M{len(rows)+1}")
    table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def main():
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f'Input workbook not found: {INPUT_XLSX}')

    wb = load_workbook(INPUT_XLSX)
    if 'S1_CohortSummary' not in wb.sheetnames:
        raise ValueError('Workbook does not contain sheet: S1_CohortSummary')

    ws = wb['S1_CohortSummary']
    existing = existing_row_map(ws)

    discovery_meta_rows = read_tsv(DISCOVERY_META)
    vel_meta_rows = read_tsv(VEL_META)
    gandal_meta_rows = read_rdata_datmeta(GANDAL_RDATA)

    disc_aud, disc_ret = load_discovery_rows(existing, discovery_meta_rows)
    val_strict, val_desc = load_validation_rows(existing, vel_meta_rows)
    gandal_row = load_gandal_row(existing, gandal_meta_rows)

    row_map = {
        (disc_aud['dataset'], disc_aud['summary_scope']): disc_aud,
        (disc_ret['dataset'], disc_ret['summary_scope']): disc_ret,
        (val_strict['dataset'], val_strict['summary_scope']): val_strict,
        (val_desc['dataset'], val_desc['summary_scope']): val_desc,
        (gandal_row['dataset'], gandal_row['summary_scope']): gandal_row,
    }
    final_rows = [row_map[key] for key in TARGET_ORDER]

    rewrite_s1_sheet(ws, final_rows)
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f'Wrote updated workbook: {OUTPUT_XLSX}')


if __name__ == '__main__':
    main()
