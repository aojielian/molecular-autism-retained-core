
from pathlib import Path
import os
import re
import gzip
import csv
from typing import List, Dict, Optional, Iterable
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(os.environ.get('MA_REVISION_ROOT', '/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/MolecularAutism_Revision'))
OUTDIR = ROOT / 'Supplementary_Tables_v2'
OUTFILE = OUTDIR / 'Supplementary_Tables_Master_v2.xlsx'
OUTDIR.mkdir(parents=True, exist_ok=True)

# -------------------------------------------------
# Manual values that are not always recoverable from exported TSVs
# -------------------------------------------------
MANUAL_S10_VALUES = {
    'discovery_clustering_resolution': '0.4',
    'discovery_reclustering_rationale': 'Selected to preserve the dominant candidate/reference contrast while avoiding over-interpretation of finer fragmentation.',
    'validation_reclustering_resolution': 'REVIEW_AND_FILL',
    'validation_reclustering_rationale': 'Fill from the final Velmeshev validation script if not recoverable from exported outputs.',
    'bulk_model_covariates': 'region, batch, sex, ancestry, age, age_squared, postmortem_interval, RNA_integrity',
    'github_repository': 'REVIEW_AND_FILL',
    'zenodo_archive': 'REVIEW_AND_FILL',
}

PACKAGE_PURPOSE = {
    'Seurat': 'single-cell preprocessing, clustering, UMAP, plotting',
    'SeuratObject': 'Seurat object infrastructure',
    'harmony': 'donor-aware integration / embedding harmonization',
    'ggplot2': 'data visualization',
    'patchwork': 'multi-panel figure assembly',
    'data.table': 'tabular data handling',
    'dplyr': 'data manipulation',
    'Matrix': 'sparse matrix operations',
    'edgeR': 'pseudobulk differential expression',
    'limma': 'linear modeling and voom workflow',
    'nlme': 'linear/mixed modeling',
    'lme4': 'linear/mixed modeling',
    'AnnotationDbi': 'gene identifier annotation',
    'org.Hs.eg.db': 'human gene annotation mapping',
    'fgsea': 'gene-set enrichment analysis',
    'clusterProfiler': 'functional enrichment / annotation support',
}

CURATED_PACKAGES = [
    'Seurat', 'SeuratObject', 'harmony', 'ggplot2', 'patchwork', 'data.table',
    'dplyr', 'Matrix', 'edgeR', 'limma', 'nlme', 'lme4', 'AnnotationDbi',
    'org.Hs.eg.db', 'fgsea', 'clusterProfiler'
]

# -------------------------------------------------
# Helpers
# -------------------------------------------------

def read_tsv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    opener = gzip.open if path.suffix == '.gz' else open
    with opener(path, 'rt', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f, delimiter='\t')
        return [dict(row) for row in reader]


def read_txt_as_rows(path: Path, value_col: str, extra: Optional[Dict[str, str]] = None) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    extra = extra or {}
    rows = []
    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.rstrip('\n')
            if line.strip() == '':
                continue
            row = {value_col: line}
            row.update(extra)
            rows.append(row)
    return rows


def maybe_int(x):
    try:
        if x is None or x == '':
            return None
        if isinstance(x, (int, float)):
            return int(x)
        return int(float(str(x)))
    except Exception:
        return None


def maybe_float(x):
    try:
        if x is None or x == '':
            return None
        return float(str(x))
    except Exception:
        return None


def normalize_str(x):
    if x is None:
        return ''
    return str(x).strip()


def clean_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    out = []
    for r in rows:
        rr = {}
        for k, v in r.items():
            k2 = normalize_str(k)
            if k2 == '' or k2 == 'source_file':
                continue
            rr[k2] = '' if v is None else str(v).strip()
        if not rr:
            continue
        if all(v == '' for v in rr.values()):
            continue
        out.append(rr)
    return out


def add_section(rows: List[Dict[str, str]], section: str) -> List[Dict[str, str]]:
    out = []
    for r in clean_rows(rows):
        rr = {'section': section}
        rr.update(r)
        out.append(rr)
    return out


def combine_rows(parts: List[List[Dict[str, str]]]) -> List[Dict[str, str]]:
    out = []
    for p in parts:
        out.extend(clean_rows(p))
    return out


def unique_values(rows: List[Dict[str, str]], candidates: List[str]) -> List[str]:
    cols = set()
    for r in rows:
        cols.update(r.keys())
    return [c for c in candidates if c in cols]


def write_sheet(ws, rows: List[Dict[str, str]]):
    rows = clean_rows(rows)
    if not rows:
        ws['A1'] = 'note'
        ws['A2'] = f'No data found for {ws.title}'
        ws.freeze_panes = 'A2'
        return

    cols = []
    seen = set()
    preferred = ['Category', 'dataset', 'summary_scope', 'analysis_unit', 'section',
                 'analysis_stage', 'cluster', 'Item', 'Version_or_value',
                 'Purpose_or_definition', 'Notes']
    for col in preferred:
        if any(col in r for r in rows) and col not in seen:
            cols.append(col)
            seen.add(col)

    for r in rows:
        for k in r.keys():
            if k not in seen:
                cols.append(k)
                seen.add(k)

    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F4E78')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='D9E2F3')
        c.border = Border(bottom=thin)

    for i, r in enumerate(rows, start=2):
        for j, col in enumerate(cols, start=1):
            ws.cell(row=i, column=j, value=r.get(col, ''))

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

    for j, col in enumerate(cols, start=1):
        max_len = len(str(col))
        for i in range(2, min(len(rows) + 2, 500)):
            val = ws.cell(row=i, column=j).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 14), 48)

    ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
    table = Table(displayName=f"Tbl_{ws.title[:20].replace('-', '_')}", ref=ref)
    style = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def find_first_existing(candidates: Iterable[Path], glob_name: Optional[str] = None) -> Optional[Path]:
    for p in candidates:
        if p.exists():
            return p
    if glob_name:
        matches = list(ROOT.rglob(glob_name))
        if matches:
            return matches[0]
    return None


def read_first_existing_tsv(candidates: Iterable[Path], glob_name: Optional[str] = None) -> List[Dict[str, str]]:
    p = find_first_existing(candidates, glob_name=glob_name)
    return read_tsv(p) if p else []


def rows_to_index(rows: List[Dict[str, str]], key: str) -> Dict[str, Dict[str, str]]:
    idx = {}
    for r in clean_rows(rows):
        k = normalize_str(r.get(key))
        if k != '':
            idx[k] = r
    return idx


def infer_background_reason(marker_preview: str) -> Optional[str]:
    s = (marker_preview or '').lower()
    if s == '':
        return None

    neuronal = ['cntnap2', 'ptprd', 'ryr2', 'camk2a', 'sv2b', 'olfm1', 'gabrb2', 'nptxr', 'st18', 'il1rapl1', 'nrg1']
    astro = ['slc1a2', 'aqp4', 'gpc5', 'ryr3']
    oligo = ['mobp', 'tf', 'slc44a1', 'opalin', 'plp1', 'mbp']
    vascular = ['colec12', 'vcan', 'f13a1', 'cr1', 'selp']
    lymphoid = ['blnk', 'dock8', 'cd83']
    stress = ['hsph1', 'dnaja4', 'bag3', 'hspd1', 'serpine1']

    if any(x in s for x in neuronal):
        return 'neuronal_or_synaptic_background'
    if any(x in s for x in astro):
        return 'astrocytic_background'
    if any(x in s for x in oligo):
        return 'oligodendroglial_or_myelin_background'
    if any(x in s for x in vascular):
        return 'vascular_or_perivascular_background'
    if any(x in s for x in lymphoid):
        return 'lymphoid_or_immune_background'
    if any(x in s for x in stress):
        return 'stress_or_uncertain_identity'
    return None


def infer_primary_reason(row: Dict[str, str], retained_clusters: set) -> str:
    cluster = normalize_str(row.get('cluster'))
    if cluster in retained_clusters:
        return 'retained_core_microglial_cluster'

    top_nonmg = maybe_float(row.get('top_nonmg_score'))
    mg_total = maybe_float(row.get('microglia_total_score'))
    if top_nonmg is not None and mg_total is not None and top_nonmg > mg_total:
        return 'non_microglial_background'

    reason_from_markers = infer_background_reason(row.get('top_marker_preview') or row.get('top50_preview'))
    if reason_from_markers:
        return reason_from_markers

    nh = maybe_int(row.get('n_homeostatic_markers_in_top50'))
    na = maybe_int(row.get('n_activation_markers_in_top50'))
    if nh == 0 and na == 0:
        return 'weak_microglial_identity'

    n_donors = maybe_int(row.get('n_donors'))
    if n_donors is not None and n_donors < 40:
        return 'limited_donor_support'

    return 'not_retained_after_audit'


def parse_session_environment(session_file: Path) -> Dict[str, str]:
    env = {}
    if not session_file or not session_file.exists():
        return env
    with open(session_file, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            s = line.strip()
            if s.startswith('R version'):
                env['R'] = s.replace('R version', '').strip()
            elif s.startswith('Platform:'):
                env['Platform'] = s.replace('Platform:', '').strip()
            elif s.startswith('Running under:'):
                env['Operating system'] = s.replace('Running under:', '').strip()
    return env


def parse_package_versions(session_files: List[Path]) -> Dict[str, str]:
    pkg_map = {}
    token_re = re.compile(r'([A-Za-z][A-Za-z0-9\.]+)_([0-9][A-Za-z0-9\.\-]+)')
    for f in session_files:
        if not f.exists():
            continue
        with open(f, 'r', encoding='utf-8', errors='ignore') as fh:
            for line in fh:
                for name, ver in token_re.findall(line):
                    if name not in pkg_map:
                        pkg_map[name] = ver
    return pkg_map


def first_metric_counts(rows: List[Dict[str, str]]) -> Dict[str, Optional[int]]:
    rows = clean_rows(rows)
    for r in rows:
        n_asd = maybe_int(r.get('n_ASD'))
        n_ctl = maybe_int(r.get('n_Control'))
        if n_asd is not None and n_ctl is not None:
            return {'n_ASD': n_asd, 'n_Control': n_ctl}
    return {'n_ASD': None, 'n_Control': None}

# -------------------------------------------------
# Package paths
# -------------------------------------------------
pkg1 = ROOT / 'Package1_Figure1_DiscoveryAudit'
pkg1b = ROOT / 'Package1b_ClusterAnnotation_ContaminantAudit'
pkg2 = ROOT / 'Package2_Define_Cluster2'
pkg2b = ROOT / 'Package2b_ResidualBackgroundAudit'
pkg3 = ROOT / 'Package3_Discovery_DonorAwareDiseaseAssociation'
pkg4b = ROOT / 'Package4b_DirectionConsistencySummary_v2_manual'
pkg4v2 = ROOT / 'Package4_CrossCohortValidation_Velmeshev_v2'
pkg5 = ROOT / 'Package5_DirectionalConcordance_GlobalSummary'
pkg7 = ROOT / 'Package7_MinimalRobustnessAnalysis'
pkg8 = ROOT / 'Package8_ExternalValidation_Gandal2022'
pkg8b = ROOT / 'Package8b_Gandal2022_LOO_Sensitivity'

# -------------------------------------------------
# S1 Cohort summary (reworked; includes discovery ASD/Control donor split)
# -------------------------------------------------
s1 = []

exact_n_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '00_Figure1_exact_n.tsv',
        pkg1 / '00_Figure1_exact_n.tsv',
    ],
    glob_name='00_Figure1_exact_n.tsv'
)

if exact_n_rows:
    for r in exact_n_rows:
        panel = normalize_str(r.get('panel'))
        diagnosis = normalize_str(r.get('diagnosis'))
        unit_rows = []
        if panel == 'A/B':
            scope = 'Discovery audited microglial object'
        elif panel == 'D/E':
            scope = 'Discovery retained-core object'
        else:
            scope = 'Discovery object'
        n_cells = maybe_int(r.get('n_cells'))
        n_donors = maybe_int(r.get('n_donors'))
        if n_donors is not None:
            unit_rows.append({
                'dataset': 'Discovery',
                'summary_scope': scope,
                'analysis_unit': 'donors',
                'n_ASD': n_donors if diagnosis == 'ASD' else '',
                'n_Control': n_donors if diagnosis == 'Control' else '',
                'n_total': n_donors if diagnosis == 'All' else '',
                'notes': f'Derived from {panel} exact n summary.'
            })
        if n_cells is not None:
            unit_rows.append({
                'dataset': 'Discovery',
                'summary_scope': scope,
                'analysis_unit': 'cells',
                'n_ASD': n_cells if diagnosis == 'ASD' else '',
                'n_Control': n_cells if diagnosis == 'Control' else '',
                'n_total': n_cells if diagnosis == 'All' else '',
                'notes': f'Derived from {panel} exact n summary.'
            })
        s1.extend(unit_rows)
else:
    rows_disc = read_tsv(pkg1 / '06_allcluster_donor_proportions.tsv.gz')
    donor_cols = unique_values(rows_disc, ['donor', 'sample', 'Subject', 'subject', 'individual_ID', 'iid'])
    dx_cols = unique_values(rows_disc, ['dx', 'Diagnosis', 'diagnosis', 'group', 'Group', 'dx_std'])
    if donor_cols and dx_cols:
        seen = {}
        for r in rows_disc:
            donor = normalize_str(r.get(donor_cols[0]))
            dx = normalize_str(r.get(dx_cols[0]))
            if donor != '' and dx != '':
                seen[donor] = dx
        n_asd = sum(v in {'ASD', 'Case', 'case', 'Autism'} for v in seen.values())
        n_ctl = sum(v in {'Control', 'CTL', 'control'} for v in seen.values())
        s1.append({
            'dataset': 'Discovery',
            'summary_scope': 'Discovery audited microglial object',
            'analysis_unit': 'donors',
            'n_ASD': n_asd,
            'n_Control': n_ctl,
            'n_total': len(seen),
            'notes': 'Fallback donor split derived from donor-proportion table.'
        })

# Validation summary
val_metric_rows = read_tsv(pkg4b / '03_direction_consistency_summary_pretty.tsv') or read_tsv(pkg4b / '04_direction_consistency_summary_manuscript.tsv')
val_counts = first_metric_counts(val_metric_rows)
if val_counts['n_ASD'] is not None or val_counts['n_Control'] is not None:
    s1.append({
        'dataset': 'Velmeshev validation',
        'summary_scope': 'Strict shared validation set',
        'analysis_unit': 'donors',
        'n_ASD': val_counts['n_ASD'] or '',
        'n_Control': val_counts['n_Control'] or '',
        'n_total': (val_counts['n_ASD'] or 0) + (val_counts['n_Control'] or 0),
        'notes': 'Strict shared donor set used for prespecified score-based validation metrics.'
    })

val_micro = read_tsv(pkg4v2 / '03_validation_microglia_dx_counts.tsv')
if val_micro and {'dx', 'N'} <= set(val_micro[0].keys()):
    n_asd_cells = sum(maybe_int(r['N']) or 0 for r in val_micro if normalize_str(r.get('dx')) == 'ASD')
    n_ctl_cells = sum(maybe_int(r['N']) or 0 for r in val_micro if normalize_str(r.get('dx')) == 'Control')
    s1.append({
        'dataset': 'Velmeshev validation',
        'summary_scope': 'Validation microglia retained for descriptive context',
        'analysis_unit': 'cells',
        'n_ASD': n_asd_cells,
        'n_Control': n_ctl_cells,
        'n_total': n_asd_cells + n_ctl_cells,
        'notes': 'Validation microglial-cell counts.'
    })

# Gandal summary
rows_gandal_scores = read_tsv(pkg8 / '07_sample_level_program_scores.tsv')
rows_gandal_dx = read_tsv(pkg8 / '03_filtered_dx_counts.tsv')
if rows_gandal_dx and {'dx', 'N'} <= set(rows_gandal_dx[0].keys()):
    n_asd = sum(maybe_int(r['N']) or 0 for r in rows_gandal_dx if normalize_str(r.get('dx')) == 'ASD')
    n_ctl = sum(maybe_int(r['N']) or 0 for r in rows_gandal_dx if normalize_str(r.get('dx')) == 'Control')
    s1.append({
        'dataset': 'Gandal 2022 bulk cortex',
        'summary_scope': 'Filtered external bulk-cortex cohort',
        'analysis_unit': 'samples',
        'n_ASD': n_asd,
        'n_Control': n_ctl,
        'n_total': n_asd + n_ctl,
        'notes': 'Filtered external bulk-cortex validation samples.'
    })
if rows_gandal_scores:
    subj_cols = unique_values(rows_gandal_scores, ['Subject', 'subject', 'donor', 'sample'])
    if subj_cols:
        subj_set = {normalize_str(r.get(subj_cols[0])) for r in rows_gandal_scores if normalize_str(r.get(subj_cols[0])) != ''}
        s1.append({
            'dataset': 'Gandal 2022 bulk cortex',
            'summary_scope': 'Filtered external bulk-cortex cohort',
            'analysis_unit': 'unique_subjects',
            'n_ASD': '',
            'n_Control': '',
            'n_total': len(subj_set),
            'notes': 'Unique subject count derived from the mapped score matrix.'
        })

# -------------------------------------------------
# S2 Discovery audit / exclusion rationale (reworked)
# -------------------------------------------------
marker_support = read_tsv(pkg1 / '10_marker_support.tsv')
cluster_priority = read_tsv(pkg1 / '11_cluster_prioritization.tsv')
top_candidate = read_tsv(pkg1 / '12_top_candidate.tsv')
cluster2_summary = read_tsv(pkg1 / '13_cluster2_summary.tsv')
panelC_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '03_panelC_underlying_values.tsv',
        pkg1 / '03_panelC_underlying_values.tsv',
    ],
    glob_name='03_panelC_underlying_values.tsv'
)
disc_cluster_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '01_discovery_clusters.tsv',
        pkg1 / '01_discovery_clusters.tsv',
    ],
    glob_name='01_discovery_clusters.tsv'
)
retained_cluster_rows = read_first_existing_tsv(
    [
        pkg1 / 'tables' / '02_retained_clusters_from_strict_object.tsv',
        pkg1 / '02_retained_clusters_from_strict_object.tsv',
    ],
    glob_name='02_retained_clusters_from_strict_object.tsv'
)

retained_clusters = set()
for r in retained_cluster_rows:
    c = normalize_str(r.get('retained_cluster') or r.get('cluster'))
    if c != '':
        retained_clusters.add(c)
if not retained_clusters:
    retained_clusters = {'0', '2'}

marker_idx = rows_to_index(marker_support, 'cluster')
priority_idx = rows_to_index(cluster_priority, 'cluster')
panelC_idx = rows_to_index(panelC_rows, 'cluster')

cluster_ids = set(marker_idx.keys()) | set(priority_idx.keys()) | set(panelC_idx.keys())
for r in disc_cluster_rows:
    c = normalize_str(r.get('cluster'))
    if c != '':
        cluster_ids.add(c)

def sort_key(x):
    try:
        return (0, int(x))
    except Exception:
        return (1, x)

s2 = []
for c in sorted(cluster_ids, key=sort_key):
    row = {'cluster': c}
    if c in marker_idx:
        row.update(marker_idx[c])
    if c in priority_idx:
        for k, v in priority_idx[c].items():
            if k not in row:
                row[k] = v
    if c in panelC_idx:
        for k, v in panelC_idx[c].items():
            if k not in row:
                row[k] = v
    row['final_decision'] = 'retained' if c in retained_clusters else 'excluded'
    row['primary_reason'] = infer_primary_reason(row, retained_clusters)
    s2.append(row)

# top-candidate metadata and cluster-2 summary appended as flat auxiliary sections
if top_candidate:
    for r in add_section(top_candidate, 'top_candidate'):
        s2.append(r)
if cluster2_summary:
    for r in add_section(cluster2_summary, 'cluster2_summary'):
        s2.append(r)

# -------------------------------------------------
# S3-S6 unchanged except cleaning
# -------------------------------------------------
s3 = combine_rows([
    add_section(read_tsv(pkg2 / 'tables' / '22_cluster2_core_summary_up.tsv'), 'candidate_program'),
    add_section(read_tsv(pkg2 / 'tables' / '23_cluster0_core_summary_up.tsv'), 'reference_program'),
])

s4 = clean_rows(read_tsv(pkg3 / '13_pseudobulk_all_retained_ASD_vs_Control.tsv.gz'))

s5 = combine_rows([
    add_section(read_tsv(pkg3 / '15_pseudobulk_cluster0_ASD_vs_Control.tsv.gz'), 'cluster0'),
    add_section(read_tsv(pkg3 / '17_pseudobulk_cluster2_ASD_vs_Control.tsv.gz'), 'cluster2'),
])

s6 = combine_rows([
    add_section(read_tsv(pkg4b / '00_shared_comparable_samples.tsv'), 'shared_comparable_samples'),
    add_section(read_tsv(pkg4b / '00b_shared_sample_counts.tsv'), 'shared_sample_counts'),
    add_section(read_tsv(pkg4b / '03_direction_consistency_summary_pretty.tsv'), 'direction_consistency_pretty'),
    add_section(read_tsv(pkg4b / '04_direction_consistency_summary_manuscript.tsv'), 'direction_consistency_manuscript'),
])

# -------------------------------------------------
# S7-S9 flattened (no blank rows / no section header blocks)
# -------------------------------------------------
s7 = combine_rows([
    add_section(read_tsv(pkg5 / '01_direction_concordance_master.tsv'), 'direction_concordance_master'),
    add_section(read_tsv(pkg5 / '02_global_direction_tests.tsv'), 'global_direction_tests'),
    add_section(read_tsv(pkg5 / '03_metrics_ranked_by_nominal_p.tsv'), 'metrics_ranked_by_nominal_p'),
    add_section(read_tsv(pkg5 / '04_metrics_ranked_by_aligned_effect.tsv'), 'metrics_ranked_by_aligned_effect'),
    add_section(read_tsv(pkg7 / '01_score_defined_harmonization_sensitivity.tsv'), 'harmonization_sensitivity'),
    add_section(read_tsv(pkg7 / '02_score_defined_direction_stability.tsv'), 'direction_stability'),
    add_section(read_tsv(pkg7 / '03_leave_one_metric_out_global_summary.tsv'), 'leave_one_metric_out_summary'),
])

s8 = combine_rows([
    add_section(read_txt_as_rows(pkg8 / '04_candidate_genes_used.txt', 'gene', {'set_name': 'candidate_genes_used'}), 'gene_membership'),
    add_section(read_txt_as_rows(pkg8 / '04_reference_genes_used.txt', 'gene', {'set_name': 'reference_genes_used'}), 'gene_membership'),
    add_section(read_txt_as_rows(pkg8 / '06_candidate_hit_genes.txt', 'gene', {'set_name': 'candidate_hit_genes'}), 'gene_membership'),
    add_section(read_txt_as_rows(pkg8 / '06_reference_hit_genes.txt', 'gene', {'set_name': 'reference_hit_genes'}), 'gene_membership'),
    add_section(read_tsv(pkg8 / '06_gene_set_overlap.tsv'), 'gene_set_overlap'),
])

s9 = combine_rows([
    add_section(read_tsv(pkg8b / '04_leave_one_region_out_results.tsv'), 'leave_one_region_out_results'),
    add_section(read_tsv(pkg8b / '05_leave_one_region_out_vs_baseline.tsv'), 'leave_one_region_out_vs_baseline'),
    add_section(read_tsv(pkg8b / '06_leave_one_region_out_stability_summary.tsv'), 'leave_one_region_out_stability_summary'),
])

# -------------------------------------------------
# S10 Structured software / package / parameter summary (reworked)
# -------------------------------------------------
session_files = [
    pkg1 / 'sessionInfo_package1.txt',
    pkg1b / 'sessionInfo_package1b.txt',
    pkg2 / 'sessionInfo_package2.txt',
    pkg2b / 'sessionInfo_package2b.txt',
    pkg4v2 / '31_sessionInfo.txt',
    pkg8 / '13_sessionInfo.txt',
    pkg8b / '11_sessionInfo.txt',
]
session_files = [p for p in session_files if p.exists()]

env_source = session_files[0] if session_files else None
env_info = parse_session_environment(env_source) if env_source else {}
pkg_versions = parse_package_versions(session_files)

s10 = []

# Environment
for item in ['R', 'Platform', 'Operating system']:
    if item in env_info:
        s10.append({
            'Category': 'Software environment',
            'analysis_stage': 'Global',
            'Item': item,
            'Version_or_value': env_info[item],
            'Purpose_or_definition': 'Core runtime environment',
            'Notes': ''
        })

# Packages
for pkg in CURATED_PACKAGES:
    ver = pkg_versions.get(pkg, '')
    if ver == '':
        continue
    s10.append({
        'Category': 'Package version',
        'analysis_stage': 'Global',
        'Item': pkg,
        'Version_or_value': ver,
        'Purpose_or_definition': PACKAGE_PURPOSE.get(pkg, ''),
        'Notes': ''
    })

# Discovery parameters
disc_all = next((r for r in exact_n_rows if normalize_str(r.get('panel')) == 'A/B' and normalize_str(r.get('diagnosis')) == 'All'), None)
disc_retained = next((r for r in exact_n_rows if normalize_str(r.get('panel')) == 'D/E' and normalize_str(r.get('diagnosis')) == 'All'), None)

disc_cluster_count = None
if disc_all:
    disc_cluster_count = maybe_int(disc_all.get('n_clusters'))
elif disc_cluster_rows:
    disc_cluster_count = len(disc_cluster_rows)

retained_cluster_count = None
if disc_retained:
    retained_cluster_count = maybe_int(disc_retained.get('n_clusters'))
elif retained_cluster_rows:
    retained_cluster_count = len(retained_cluster_rows)

s10.extend([
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Discovery clustering resolution',
        'Version_or_value': MANUAL_S10_VALUES['discovery_clustering_resolution'],
        'Purpose_or_definition': 'Resolution used to define the reclustered discovery microglial analysis space.',
        'Notes': MANUAL_S10_VALUES['discovery_reclustering_rationale'],
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Discovery clusters',
        'Version_or_value': disc_cluster_count if disc_cluster_count is not None else '',
        'Purpose_or_definition': 'Number of clusters in the audited discovery-stage microglial analysis space.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Retained clusters',
        'Version_or_value': ', '.join(sorted(retained_clusters, key=sort_key)),
        'Purpose_or_definition': 'Clusters retained as the cortical microglial core for downstream inference.',
        'Notes': '',
    },
])

if disc_all:
    s10.append({
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Discovery audited microglial object',
        'Version_or_value': f"{disc_all.get('n_cells')} cells from {disc_all.get('n_donors')} donors",
        'Purpose_or_definition': 'Audited discovery-stage microglial analysis object.',
        'Notes': '',
    })
if disc_retained:
    s10.append({
        'Category': 'Analysis parameter',
        'analysis_stage': 'Discovery',
        'Item': 'Strict retained-core object',
        'Version_or_value': f"{disc_retained.get('n_cells')} cells from {disc_retained.get('n_donors')} donors",
        'Purpose_or_definition': 'Strict retained-core microglial object used for downstream analyses.',
        'Notes': '',
    })

# Validation parameters
s10.append({
    'Category': 'Analysis parameter',
    'analysis_stage': 'Validation',
    'Item': 'Validation reclustering resolution',
    'Version_or_value': MANUAL_S10_VALUES['validation_reclustering_resolution'],
    'Purpose_or_definition': 'Resolution used for validation microglia reclustering in the Velmeshev dataset.',
    'Notes': MANUAL_S10_VALUES['validation_reclustering_rationale'],
})

if val_counts['n_ASD'] is not None or val_counts['n_Control'] is not None:
    s10.append({
        'Category': 'Analysis parameter',
        'analysis_stage': 'Validation',
        'Item': 'Strict shared validation set',
        'Version_or_value': f"ASD={val_counts['n_ASD']}; Control={val_counts['n_Control']}",
        'Purpose_or_definition': 'Shared donor set used for the seven prespecified score-based validation metrics.',
        'Notes': '',
    })

global_dir_tests = read_tsv(pkg5 / '02_global_direction_tests.tsv')
n_metrics = ''
for r in global_dir_tests:
    if normalize_str(r.get('test_name')) == 'n_total_metrics':
        n_metrics = r.get('value', '')
        break
s10.append({
    'Category': 'Analysis parameter',
    'analysis_stage': 'Validation',
    'Item': 'Prespecified score-based validation metrics',
    'Version_or_value': n_metrics,
    'Purpose_or_definition': 'Number of primary score-based validation readouts used for directional-concordance testing.',
    'Notes': '',
})

# Bulk parameters
s10.extend([
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'External bulk cortex',
        'Item': 'External bulk validation dataset',
        'Version_or_value': 'Gandal 2022 bulk cortex',
        'Purpose_or_definition': 'Independent tissue-level validation cohort.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'External bulk cortex',
        'Item': 'Primary external scores',
        'Version_or_value': 'candidate score; candidate-minus-reference score; reference score',
        'Purpose_or_definition': 'Bulk projection readouts derived from the frozen candidate/reference programs.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'External bulk cortex',
        'Item': 'Bulk model covariates',
        'Version_or_value': MANUAL_S10_VALUES['bulk_model_covariates'],
        'Purpose_or_definition': 'Covariates included in the mixed-model ASD effect testing framework.',
        'Notes': '',
    },
    {
        'Category': 'Analysis parameter',
        'analysis_stage': 'External bulk cortex',
        'Item': 'Sensitivity analysis',
        'Version_or_value': 'leave-one-region-out',
        'Purpose_or_definition': 'Region-wise robustness analysis for the external bulk-cortex validation.',
        'Notes': '',
    },
])

# Reproducibility
s10.extend([
    {
        'Category': 'Reproducibility resource',
        'analysis_stage': 'Global',
        'Item': 'GitHub repository',
        'Version_or_value': MANUAL_S10_VALUES['github_repository'],
        'Purpose_or_definition': 'Analysis scripts and figure-generation code.',
        'Notes': 'Replace with the final public repository URL before submission.',
    },
    {
        'Category': 'Reproducibility resource',
        'analysis_stage': 'Global',
        'Item': 'Zenodo archive',
        'Version_or_value': MANUAL_S10_VALUES['zenodo_archive'],
        'Purpose_or_definition': 'Versioned archival release linked to the submitted manuscript.',
        'Notes': 'Replace with the final DOI or persistent archive URL before submission.',
    },
])

# -------------------------------------------------
# Write workbook
# -------------------------------------------------
wb = Workbook()
wb.remove(wb.active)

sheet_map = [
    ('S1_CohortSummary', s1),
    ('S2_DiscoveryAudit', s2),
    ('S3_GenePrograms', s3),
    ('S4_PooledDEGs', s4),
    ('S5_ClusterDEGs', s5),
    ('S6_ValidationMetrics', s6),
    ('S7_ConcordanceRobust', s7),
    ('S8_GandalMapping', s8),
    ('S9_LOO_Results', s9),
    ('S10_SessionParams', s10),
]

for name, rows in sheet_map:
    ws = wb.create_sheet(title=name[:31])
    write_sheet(ws, rows)

wb.save(OUTFILE)
print(f'Wrote workbook: {OUTFILE}')
