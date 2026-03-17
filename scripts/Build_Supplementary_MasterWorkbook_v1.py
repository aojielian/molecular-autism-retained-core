from pathlib import Path
import gzip
import csv
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path('/gpfs/hpc/home/lijc/lianaoj/autism_scRNA/MolecularAutism_Revision')
OUTDIR = ROOT / 'Supplementary_Tables_v1'
OUTFILE = OUTDIR / 'Supplementary_Tables_Master.xlsx'
OUTDIR.mkdir(parents=True, exist_ok=True)

# ------------------------------
# Helpers
# ------------------------------

def read_tsv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    opener = gzip.open if path.suffix == '.gz' else open
    mode = 'rt'
    with opener(path, mode, encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f, delimiter='\t')
        return [dict(row) for row in reader]


def read_txt_as_rows(path: Path, value_col: str, extra: Optional[Dict[str, str]] = None) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    extra = extra or {}
    rows = []
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.rstrip('\n')
            if line == '':
                continue
            row = {value_col: line}
            row.update(extra)
            rows.append(row)
    return rows


def unique_values(rows: List[Dict[str, str]], candidates: List[str]) -> List[str]:
    cols = set()
    for r in rows:
        cols.update(r.keys())
    return [c for c in candidates if c in cols]


def maybe_int(x):
    try:
        if x is None or x == '':
            return None
        if isinstance(x, (int, float)):
            return int(x)
        return int(float(str(x)))
    except Exception:
        return None


def add_section(rows: List[Dict[str, str]], section: str) -> List[Dict[str, str]]:
    out = []
    for r in rows:
        rr = {'section': section}
        rr.update(r)
        out.append(rr)
    return out


def drop_source_file(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    out = []
    for r in rows:
        rr = {k: v for k, v in r.items() if k != 'source_file'}
        out.append(rr)
    return out


def combine_rows(parts: List[List[Dict[str, str]]]) -> List[Dict[str, str]]:
    out = []
    for p in parts:
        out.extend(p)
    return drop_source_file(out)


def write_sheet(ws, rows: List[Dict[str, str]], title: str):
    # If empty, still write title row with a note
    if not rows:
        ws['A1'] = 'note'
        ws['A2'] = f'No data found for {title}'
        ws.freeze_panes = 'A2'
        return

    # Determine column order: section first if present, then others in first-seen order
    cols = []
    seen = set()
    if any('section' in r for r in rows):
        cols.append('section')
        seen.add('section')
    for r in rows:
        for k in r.keys():
            if k not in seen and k != 'source_file':
                cols.append(k)
                seen.add(k)

    # Write header
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F4E78')
        c.alignment = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='D9E2F3')
        c.border = Border(bottom=thin)

    # Write rows
    for i, r in enumerate(rows, start=2):
        for j, col in enumerate(cols, start=1):
            val = r.get(col, '')
            ws.cell(row=i, column=j, value=val)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(2, len(rows)+1)}"

    # Widths
    for j, col in enumerate(cols, start=1):
        max_len = len(str(col))
        for i in range(2, min(len(rows)+2, 300)):  # enough for sizing without too much work
            val = ws.cell(row=i, column=j).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 12), 42)

    # Add table style if there is at least one data row
    if len(rows) >= 1:
        ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
        table = Table(displayName=f"Tbl_{ws.title[:20].replace('-', '_')}", ref=ref)
        style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

# ------------------------------
# Build rows for each supplementary table
# ------------------------------

pkg1 = ROOT / 'Package1_Figure1_DiscoveryAudit'
pkg1b = ROOT / 'Package1b_ClusterAnnotation_ContaminantAudit'
pkg2 = ROOT / 'Package2_Define_Cluster2'
pkg3 = ROOT / 'Package3_Discovery_DonorAwareDiseaseAssociation'
pkg4b = ROOT / 'Package4b_DirectionConsistencySummary_v2_manual'
pkg4v2 = ROOT / 'Package4_CrossCohortValidation_Velmeshev_v2'
pkg5 = ROOT / 'Package5_DirectionalConcordance_GlobalSummary'
pkg7 = ROOT / 'Package7_MinimalRobustnessAnalysis'
pkg8 = ROOT / 'Package8_ExternalValidation_Gandal2022'
pkg8b = ROOT / 'Package8b_Gandal2022_LOO_Sensitivity'

# S1 Cohort summary
s1 = []
# Discovery
rows_disc = read_tsv(pkg1 / '06_allcluster_donor_proportions.tsv.gz')
rows_disc_meta = read_tsv(pkg1 / 'Package1_Audit_Metadata.tsv')
row = {'cohort': 'Discovery', 'source_package': 'Package1_Figure1_DiscoveryAudit'}
donor_cols = unique_values(rows_disc, ['donor', 'sample', 'Subject', 'subject', 'individual_ID', 'iid'])
dx_cols = unique_values(rows_disc, ['dx', 'Diagnosis', 'diagnosis', 'group', 'Group', 'dx_std'])
if donor_cols:
    row['n_donors'] = len({r.get(donor_cols[0], '') for r in rows_disc if r.get(donor_cols[0], '')})
if dx_cols:
    vals = [r.get(dx_cols[0], '') for r in rows_disc]
    row['n_ASD'] = sum(v in {'ASD', 'Case', 'case', 'Autism'} for v in vals)
    row['n_Control'] = sum(v in {'Control', 'CTL', 'control'} for v in vals)
if rows_disc_meta:
    meta_cols = set().union(*[set(r.keys()) for r in rows_disc_meta])
    if 'n_cells' in meta_cols:
        row['n_cells'] = sum(maybe_int(r.get('n_cells')) or 0 for r in rows_disc_meta)
    if 'n_samples' in meta_cols:
        row['n_samples'] = sum(maybe_int(r.get('n_samples')) or 0 for r in rows_disc_meta)
row['notes'] = 'Discovery cohort summary assembled from audit outputs.'
s1.append(row)
# Validation
val_global = read_tsv(pkg4v2 / '02_validation_global_dx_counts.tsv')
val_micro = read_tsv(pkg4v2 / '03_validation_microglia_dx_counts.tsv')
row = {'cohort': 'Velmeshev validation', 'source_package': 'Package4_CrossCohortValidation_Velmeshev_v2'}
if val_global and {'dx','N'} <= set(val_global[0].keys()):
    row['n_ASD'] = sum(maybe_int(r['N']) or 0 for r in val_global if r.get('dx') == 'ASD')
    row['n_Control'] = sum(maybe_int(r['N']) or 0 for r in val_global if r.get('dx') == 'Control')
if val_micro and {'dx','N'} <= set(val_micro[0].keys()):
    row['n_cells'] = sum(maybe_int(r['N']) or 0 for r in val_micro)
row['notes'] = 'Validation microglial-cell counts.'
s1.append(row)
# Gandal
rows_gandal_scores = read_tsv(pkg8 / '07_sample_level_program_scores.tsv')
rows_gandal_dx = read_tsv(pkg8 / '03_filtered_dx_counts.tsv')
row = {'cohort': 'Gandal 2022 bulk cortex', 'source_package': 'Package8_ExternalValidation_Gandal2022'}
if rows_gandal_scores:
    row['n_samples'] = len(rows_gandal_scores)
    subj_cols = unique_values(rows_gandal_scores, ['Subject', 'subject', 'donor', 'sample'])
    if subj_cols:
        row['n_donors'] = len({r.get(subj_cols[0], '') for r in rows_gandal_scores if r.get(subj_cols[0], '')})
if rows_gandal_dx and {'dx','N'} <= set(rows_gandal_dx[0].keys()):
    row['n_ASD'] = sum(maybe_int(r['N']) or 0 for r in rows_gandal_dx if r.get('dx') == 'ASD')
    row['n_Control'] = sum(maybe_int(r['N']) or 0 for r in rows_gandal_dx if r.get('dx') == 'Control')
row['notes'] = 'External bulk-cortex validation cohort.'
s1.append(row)

# S2 Discovery audit / exclusion rationale
s2 = combine_rows([
    add_section(read_tsv(pkg1 / '10_marker_support.tsv'), 'marker_support'),
    add_section(read_tsv(pkg1 / '11_cluster_prioritization.tsv'), 'cluster_prioritization'),
    add_section(read_tsv(pkg1 / '12_top_candidate.tsv'), 'top_candidate'),
    add_section(read_tsv(pkg1 / '13_cluster2_summary.tsv'), 'cluster2_summary')
])

# S3 Frozen gene programs
s3 = combine_rows([
    add_section(read_tsv(pkg2 / 'tables' / '22_cluster2_core_summary_up.tsv'), 'candidate_cluster2'),
    add_section(read_tsv(pkg2 / 'tables' / '23_cluster0_core_summary_up.tsv'), 'reference_cluster0')
])

# S4 pooled DEGs
s4 = read_tsv(pkg3 / '13_pseudobulk_all_retained_ASD_vs_Control.tsv.gz')

# S5 cluster DEGs
s5 = combine_rows([
    add_section(read_tsv(pkg3 / '15_pseudobulk_cluster0_ASD_vs_Control.tsv.gz'), 'cluster0'),
    add_section(read_tsv(pkg3 / '17_pseudobulk_cluster2_ASD_vs_Control.tsv.gz'), 'cluster2')
])

# S6 validation metric summary
s6 = combine_rows([
    add_section(read_tsv(pkg4b / '00_shared_comparable_samples.tsv'), 'shared_comparable_samples'),
    add_section(read_tsv(pkg4b / '00b_shared_sample_counts.tsv'), 'shared_sample_counts'),
    add_section(read_tsv(pkg4b / '03_direction_consistency_summary_pretty.tsv'), 'direction_consistency_pretty'),
    add_section(read_tsv(pkg4b / '04_direction_consistency_summary_manuscript.tsv'), 'direction_consistency_manuscript')
])

# S7 concordance + robustness
s7 = combine_rows([
    add_section(read_tsv(pkg5 / '01_direction_concordance_master.tsv'), 'direction_concordance_master'),
    add_section(read_tsv(pkg5 / '02_global_direction_tests.tsv'), 'global_direction_tests'),
    add_section(read_tsv(pkg5 / '03_metrics_ranked_by_nominal_p.tsv'), 'metrics_ranked_by_nominal_p'),
    add_section(read_tsv(pkg5 / '04_metrics_ranked_by_aligned_effect.tsv'), 'metrics_ranked_by_aligned_effect'),
    add_section(read_tsv(pkg7 / '01_score_defined_harmonization_sensitivity.tsv'), 'harmonization_sensitivity'),
    add_section(read_tsv(pkg7 / '02_score_defined_direction_stability.tsv'), 'direction_stability'),
    add_section(read_tsv(pkg7 / '03_leave_one_metric_out_global_summary.tsv'), 'leave_one_metric_out_summary')
])

# S8 Gandal mapping summary
s8 = combine_rows([
    read_txt_as_rows(pkg8 / '04_candidate_genes_used.txt', 'gene', {'set_name': 'candidate_genes_used'}),
    read_txt_as_rows(pkg8 / '04_reference_genes_used.txt', 'gene', {'set_name': 'reference_genes_used'}),
    read_txt_as_rows(pkg8 / '06_candidate_hit_genes.txt', 'gene', {'set_name': 'candidate_hit_genes'}),
    read_txt_as_rows(pkg8 / '06_reference_hit_genes.txt', 'gene', {'set_name': 'reference_hit_genes'}),
    add_section(read_tsv(pkg8 / '06_gene_set_overlap.tsv'), 'gene_set_overlap')
])

# S9 leave-one-region-out
s9 = combine_rows([
    add_section(read_tsv(pkg8b / '04_leave_one_region_out_results.tsv'), 'leave_one_region_out_results'),
    add_section(read_tsv(pkg8b / '05_leave_one_region_out_vs_baseline.tsv'), 'leave_one_region_out_vs_baseline'),
    add_section(read_tsv(pkg8b / '06_leave_one_region_out_stability_summary.tsv'), 'leave_one_region_out_stability_summary')
])

# S10 session and parameter summary
session_files = [
    pkg1 / 'sessionInfo_package1.txt',
    pkg1b / 'sessionInfo_package1b.txt',
    pkg2 / 'sessionInfo_package2.txt',
    ROOT / 'Package2b_ResidualBackgroundAudit' / 'sessionInfo_package2b.txt',
    pkg4v2 / '31_sessionInfo.txt',
    pkg5 / '06_manuscript_summary.txt',
    pkg7 / '04_package7_summary.txt',
    pkg8 / '13_sessionInfo.txt',
    pkg8b / '11_sessionInfo.txt'
]
s10 = []
for f in session_files:
    if f.exists():
        with open(f, 'r', encoding='utf-8', errors='ignore') as fh:
            for i, line in enumerate(fh, start=1):
                s10.append({'file': f.name, 'line_no': i, 'text': line.rstrip('\n')})

# ------------------------------
# Write workbook
# ------------------------------
wb = Workbook()
def_sheet = wb.active
wb.remove(def_sheet)

sheet_map = [
    ('S1_CohortSummary', s1),
    ('S2_DiscoveryAudit', s2),
    ('S3_GenePrograms', s3),
    ('S4_PooledDEGs', s4),
    ('S5_ClusterDEGs', s5),
    ('S6_ValidationMetrics', s6),
    ('S7_ConcordanceRobust', s7),
    ('S8_GandalMapping', s8),
    ('S9_LOO_Results', s9),
    ('S10_SessionParams', s10),
]

for name, rows in sheet_map:
    ws = wb.create_sheet(title=name[:31])
    write_sheet(ws, rows, name)

wb.save(OUTFILE)
print(f'Wrote workbook: {OUTFILE}')

